import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import os
import io
import json
import hashlib
import secrets
import warnings
warnings.filterwarnings('ignore')

# ================================a+============================
# WOODMAT — ROTATION DU STOCK (Web App)
# ============================================================

st.set_page_config(page_title="WOODMAT — Rotation du stock", layout="wide",
                    page_icon="📦", initial_sidebar_state="expanded")

APP_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_HISTORIQUE = os.path.join(APP_DIR, "base_mouvements.pkl")  # base 2020-2025, livrée avec l'app
USERS_FILE = os.path.join(APP_DIR, "woodmat_users.json")
PARAMS_FILE = os.path.join(APP_DIR, "parametres_stock.json")
SEUIL = 0.001

# ── Paramètres de réapprovisionnement — valeurs de repli utilisées UNIQUEMENT si
#    aucun paramètre n'a encore été saisi pour une catégorie dans la page Paramètres.
#    Jamais déduites des données : ce sont des constantes documentées. ──
DEFAULT_PARAMS_CATEGORIE = {
    'lead_time_mois': 1.0,      # Délai d'approvisionnement (mois)
    'seuil_rupture': 1.0,       # Seuil de rupture (mois)
    'stock_securite': 0.0,      # Stock de sécurité (mois)
    'stock_cible': 3.0,         # Stock cible (mois)
}

CLASS_COLORS = {
    'Rupture': '#FFC7CE', 'Critique': '#FF8A8A', 'Stock faible': '#FFD9A0',
    'Normal': '#FFEB9C', 'Bon niveau': '#C6EFCE', 'Surstock': '#BDD7EE',
    'Surstock important': '#9FA8B5', 'Dormant': '#F4B942', 'Sans mouvement': '#E0E0E0',
}

# ============================================================
# AUTHENTIFICATION
# ============================================================

def inject_global_styles():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600;9..144,700&family=Inter:wght@400;500;600;700&family=IBM+Plex+Mono:wght@500;600;700&display=swap');

        :root {
            --wm-paper: #EDE7DA;
            --wm-panel: #F8F5EE;
            --wm-ink: #241C13;
            --wm-ink-soft: #6B5F4D;
            --wm-brass: #A6672B;
            --wm-brass-dark: #7D4C1E;
            --wm-steel: #2B4A5C;
            --wm-steel-dark: #182F3B;
            --wm-line: rgba(36, 28, 19, 0.14);
            --wm-shadow: rgba(36, 28, 19, 0.10);
            --woodmat-blue: var(--wm-steel);
            --woodmat-gold: var(--wm-brass);
            --woodmat-border: var(--wm-line);
        }

        html, body, .stApp { background: var(--wm-paper); color: var(--wm-ink); }
        .stApp {
            background-image:
                repeating-linear-gradient(115deg, rgba(36,28,19,0.025) 0px, rgba(36,28,19,0.025) 1px, transparent 1px, transparent 34px);
        }
        * { font-family: 'Inter', sans-serif; }
        h1, h2, h3, .woodmat-page-title, .woodmat-header-title, .woodmat-kpi-title, .woodmat-section-title {
            font-family: 'Fraunces', serif !important; letter-spacing: -0.01em;
        }
        .woodmat-kpi-value, [data-testid="stMetricValue"], .stDataFrame, [data-testid="stDataFrame"] * {
            font-variant-numeric: tabular-nums;
        }
        .woodmat-kpi-value { font-family: 'IBM Plex Mono', monospace !important; letter-spacing: -0.02em; }

        [data-testid="stSidebar"] {
            background: linear-gradient(195deg, var(--wm-steel-dark) 0%, var(--wm-steel) 100%);
            border-right: 1px solid rgba(0,0,0,0.25);
        }
        [data-testid="stSidebar"] * { color: #F3EFE6 !important; }
        [data-testid="stSidebar"] .stRadio label {
            border-radius: 10px; padding: 0.4rem 0.6rem; transition: background 0.15s ease;
        }
        [data-testid="stSidebar"] .stRadio label:hover { background: rgba(243,239,230,0.08); }
        [data-testid="stSidebar"] .stRadio [data-baseweb="radio"] div:first-child { border-color: var(--wm-brass) !important; }
        [data-testid="stSidebar"] hr { border-color: rgba(243,239,230,0.16); }
        [data-testid="stSidebar"] .stButton > button, [data-testid="stSidebar"] button {
            border-radius: 10px; border: 1px solid rgba(243,239,230,0.25) !important;
            background: rgba(243,239,230,0.06) !important;
        }
        [data-testid="stSidebar"] .stButton > button:hover { background: var(--wm-brass) !important; border-color: var(--wm-brass) !important; }

        .block-container { padding-top: 1.25rem; padding-bottom: 2rem; }

        /* Boutons, champs, onglets natifs Streamlit */
        .stButton > button {
            border-radius: 8px; border: 1px solid var(--wm-line); background: var(--wm-panel);
            color: var(--wm-ink); font-weight: 600; transition: all 0.15s ease;
        }
        .stButton > button:hover { border-color: var(--wm-brass); color: var(--wm-brass-dark); }
        .stButton > button[kind="primary"] {
            background: var(--wm-brass); border-color: var(--wm-brass-dark); color: #FBF8F2;
        }
        .stButton > button[kind="primary"]:hover { background: var(--wm-brass-dark); }

        [data-testid="stTabs"] [data-baseweb="tab-list"] {
            gap: 4px; border-bottom: 2px solid var(--wm-line);
        }
        [data-testid="stTabs"] [data-baseweb="tab"] {
            border-radius: 10px 10px 0 0; padding: 0.55rem 1rem; font-weight: 600; color: var(--wm-ink-soft);
        }
        [data-testid="stTabs"] [aria-selected="true"] {
            background: var(--wm-panel); color: var(--wm-brass-dark) !important;
            border: 1px solid var(--wm-line); border-bottom: 2px solid var(--wm-panel);
        }

        [data-testid="stDataFrame"] { border: 1px solid var(--wm-line); border-radius: 12px; overflow: hidden; }
        [data-testid="stDataFrame"] [role="columnheader"] {
            background: var(--wm-steel) !important; color: #F3EFE6 !important; font-weight: 600 !important;
        }

        [data-testid="stMetric"] {
            background: var(--wm-panel); border: 1px solid var(--wm-line); border-radius: 14px;
            padding: 0.8rem 1rem; box-shadow: 0 6px 18px var(--wm-shadow);
        }
        [data-testid="stMetricValue"] { color: var(--wm-ink); font-family: 'IBM Plex Mono', monospace; }
        [data-testid="stMetricLabel"] { color: var(--wm-ink-soft); }

        [data-testid="stExpander"] { border: 1px solid var(--wm-line); border-radius: 12px; background: var(--wm-panel); }

        input, textarea, select, .stSelectbox div, .stMultiSelect div, .stNumberInput input {
            border-radius: 8px !important;
        }
        .stSlider [data-baseweb="slider"] div[role="slider"] { background-color: var(--wm-brass) !important; }

        /* Éléments maison WOODMAT */
        .woodmat-login-shell {
            max-width: 460px; margin: 7vh auto 0; padding: 2.2rem;
            background: var(--wm-panel); border: 1px solid var(--wm-line);
            border-radius: 16px; box-shadow: 0 20px 60px var(--wm-shadow);
        }
        .woodmat-logo {
            width: 64px; height: 64px; border-radius: 12px; margin: 0 auto 0.8rem;
            display: flex; align-items: center; justify-content: center;
            background: linear-gradient(135deg, var(--wm-steel), var(--wm-brass));
            color: #FBF8F2; font-size: 1.7rem; font-weight: 700; font-family: 'Fraunces', serif;
            box-shadow: inset 0 0 0 3px rgba(251,248,242,0.18);
        }
        .woodmat-header {
            position: sticky; top: 0; z-index: 50; margin-bottom: 1.2rem;
            display: flex; align-items: center; justify-content: space-between; gap: 1rem;
            padding: 0.85rem 1.1rem; background: rgba(248,245,238,0.94);
            border: 1px solid var(--wm-line); border-left: 4px solid var(--wm-brass); border-radius: 12px;
            box-shadow: 0 8px 24px var(--wm-shadow); backdrop-filter: blur(8px);
        }
        .woodmat-header-title { color: var(--wm-ink); font-weight: 700; font-size: 1.12rem; }
        .woodmat-header-meta { color: var(--wm-ink-soft); font-size: 0.87rem; text-align: right; font-family: 'IBM Plex Mono', monospace; }
        .woodmat-page-title {
            color: var(--wm-ink); margin: 0 0 0.35rem; font-weight: 700; position: relative;
            padding-left: 0.85rem; border-left: 4px solid var(--wm-brass);
        }
        .woodmat-panel, .woodmat-kpi-card {
            background: var(--wm-panel); border: 1px solid var(--wm-line);
            border-radius: 14px; padding: 1rem 1.1rem;
            box-shadow: 0 6px 18px var(--wm-shadow);
        }
        .woodmat-kpi-card {
            min-height: 150px; position: relative; border-top: 3px solid var(--wm-brass);
        }
        .woodmat-kpi-title {
            color: var(--wm-ink-soft); font-size: 0.82rem; font-weight: 600; margin-bottom: 0.5rem;
            text-transform: uppercase; letter-spacing: 0.06em; font-family: 'Inter', sans-serif !important;
        }
        .woodmat-kpi-value { color: var(--wm-ink); font-size: 1.5rem; font-weight: 700; line-height: 1.3; }
        .woodmat-kpi-detail { color: var(--wm-ink-soft); font-size: 0.93rem; line-height: 1.45; margin-top: 0.15rem; }
        .woodmat-muted { color: var(--wm-ink-soft); font-size: 0.9rem; line-height: 1.35; margin-top: 0.35rem; }
        .woodmat-section-title {
            color: var(--wm-ink); font-size: 1.1rem; font-weight: 700; margin-bottom: 0.1rem;
        }
        .woodmat-legend { color: var(--wm-ink-soft); font-size: 0.86rem; line-height: 1.45; margin-top: 0.35rem; }
        .woodmat-coming-soon {
            border: 1px dashed var(--wm-ink-soft); border-radius: 14px; padding: 1.2rem;
            background: rgba(248,245,238,0.7); color: var(--wm-ink-soft);
        }
        @media (max-width: 768px) { .woodmat-header { flex-direction: column; align-items: flex-start; } .woodmat-header-meta { text-align: left; } }
        </style>
        """,
        unsafe_allow_html=True)


def hash_password(password, salt=None):
    """Hash PBKDF2 local pour les comptes applicatifs."""
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 120_000)
    return f"{salt}${digest.hex()}"


def verify_password(password, stored_hash):
    try:
        salt, _digest = stored_hash.split('$', 1)
    except ValueError:
        return False
    return secrets.compare_digest(hash_password(password, salt), stored_hash)


def default_users():
    now = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')
    return [{
        "name": "Administrateur WOODMAT", "email": "admin@woodmat.local",
        "password_hash": hash_password("woodmat2026"), "role": "Administrateur",
        "created_at": now, "last_login": "—", "status": "Actif"
    }]


def load_users():
    if not os.path.exists(USERS_FILE):
        users = default_users()
        save_users(users)
        return users
    try:
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return default_users()


def save_users(users):
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(users, f, ensure_ascii=False, indent=2)


def get_current_user():
    email = st.session_state.get("auth_email")
    for u in load_users():
        if u.get("email") == email:
            return u
    return st.session_state.get("auth_user", {})


def user_can_manage_users():
    return get_current_user().get("role") == "Administrateur"


def check_login():
    if st.session_state.get("auth_ok"):
        return True

    inject_global_styles()
    st.markdown(
        """
        <div class='woodmat-login-shell'>
            <div class='woodmat-logo'>W</div>
            <div style='text-align:center'>
                <h1 style='color:#1F3864;margin-bottom:0'>WOODMAT</h1>
                <p style='color:#64748B;margin-top:0.35rem'>Application métier — gestion et analyse des stocks</p>
            </div>
        </div>
        """,
        unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.15, 1])
    with col2:
        with st.form("login_form"):
            email = st.text_input("Email", placeholder="admin@woodmat.local", key="login_email")
            pwd = st.text_input("Mot de passe", type="password", key="login_password")
            submit = st.form_submit_button("Se connecter", use_container_width=True)
        st.caption("Compte initial : admin@woodmat.local / woodmat2026 — à modifier après la première connexion.")
        if submit:
            users = load_users()
            user = next((u for u in users if u.get("email", "").lower() == email.lower().strip()), None)
            if user and user.get("status") == "Actif" and verify_password(pwd, user.get("password_hash", "")):
                last_login = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')
                user["last_login"] = last_login
                save_users(users)
                st.session_state["auth_ok"] = True
                st.session_state["auth_email"] = user["email"]
                st.session_state["auth_user"] = user
                st.session_state["last_login"] = last_login
                st.rerun()
            elif user and user.get("status") != "Actif":
                st.error("Compte désactivé. Contactez votre administrateur WOODMAT.")
            else:
                st.error("Identifiants incorrects.")
    return False


def render_user_management():
    st.markdown("<h2 class='woodmat-page-title'>👥 Gestion des utilisateurs</h2>", unsafe_allow_html=True)
    if not user_can_manage_users():
        st.error("Accès réservé à l'Administrateur.")
        st.stop()
    users = load_users()
    st.caption("Création, modification, désactivation et réinitialisation des accès WOODMAT.")
    with st.expander("➕ Ajouter un utilisateur", expanded=True):
        with st.form("add_user"):
            c1, c2 = st.columns(2)
            name = c1.text_input("Nom", key="admin_add_user_nom")
            email = c2.text_input("Email", key="admin_add_user_email")
            role = c1.selectbox("Rôle", ["Direction", "Commercial", "Administrateur"], key="admin_add_user_role")
            pwd = c2.text_input("Mot de passe initial", type="password", key="admin_add_user_password")
            if st.form_submit_button("Créer l'utilisateur", type="primary"):
                if not name or not email or not pwd:
                    st.error("Nom, email et mot de passe sont obligatoires.")
                elif any(u.get("email", "").lower() == email.lower().strip() for u in users):
                    st.error("Cet email existe déjà.")
                else:
                    users.append({"name": name, "email": email.lower().strip(), "password_hash": hash_password(pwd),
                                  "role": role, "created_at": pd.Timestamp.now().strftime('%d/%m/%Y %H:%M'),
                                  "last_login": "—", "status": "Actif"})
                    save_users(users); st.success("Utilisateur créé."); st.rerun()

    display = pd.DataFrame([{k: u.get(k) for k in ["name", "email", "role", "created_at", "last_login", "status"]} for u in users])
    st.dataframe(display.rename(columns={"name":"Nom", "email":"Email", "role":"Rôle", "created_at":"Date de création", "last_login":"Dernière connexion", "status":"Statut"}), use_container_width=True)
    selected = st.selectbox("Utilisateur à administrer", [u["email"] for u in users], key="admin_select_user")
    user = next(u for u in users if u["email"] == selected)
    with st.form("edit_user"):
        c1, c2, c3 = st.columns(3)
        new_name = c1.text_input("Nom", value=user.get("name", ""), key="admin_edit_user_nom")
        new_role = c2.selectbox("Rôle", ["Administrateur", "Direction", "Commercial"], index=["Administrateur", "Direction", "Commercial"].index(user.get("role", "Direction")), key="admin_edit_user_role")
        new_status = c3.selectbox("Statut", ["Actif", "Désactivé"], index=0 if user.get("status") == "Actif" else 1, key="admin_edit_user_statut")
        new_pwd = st.text_input("Nouveau mot de passe (laisser vide pour ne pas changer)", type="password", key="admin_edit_user_password")
        a, b, c = st.columns(3)
        save_btn = a.form_submit_button("Modifier")
        reset_btn = b.form_submit_button("Réinitialiser le mot de passe")
        delete_btn = c.form_submit_button("Supprimer")
        if save_btn or reset_btn or delete_btn:
            if delete_btn:
                if user["email"] == st.session_state.get("auth_email"):
                    st.error("Vous ne pouvez pas supprimer votre propre compte connecté.")
                else:
                    users = [u for u in users if u["email"] != selected]
                    save_users(users); st.success("Utilisateur supprimé."); st.rerun()
            else:
                user["name"], user["role"], user["status"] = new_name, new_role, new_status
                if new_pwd:
                    user["password_hash"] = hash_password(new_pwd)
                if reset_btn and not new_pwd:
                    st.warning("Saisissez un nouveau mot de passe avant de réinitialiser.")
                else:
                    save_users(users); st.success("Utilisateur mis à jour."); st.rerun()


# ============================================================
# CHARGEMENT / CALCUL — vectorisé pour performance (60 000+ lignes)
# ============================================================

def parse_qty_series(s):
    """Nettoyage + conversion numérique vectorisée (rapide sur gros volumes)."""
    s = s.astype(str).str.strip()
    s = (s.str.replace(' ', '', regex=False)
          .str.replace(',', '.', regex=False)
          .str.replace('-', '', regex=False)
          .str.replace('+', '', regex=False))
    return pd.to_numeric(s, errors='coerce').fillna(0.0)


def parse_dates_series(s):
    d = pd.to_datetime(s, format='%d-%m-%y %H:%M:%S', errors='coerce')
    mask = d.isna()
    if mask.any():
        d.loc[mask] = pd.to_datetime(s[mask], dayfirst=True, errors='coerce')
    return d


def normaliser_mouv(df):
    if not str(df.columns[0]).upper().startswith('DATE'):
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)

    ncols = len(df.columns)
    if ncols >= 14:
        df.columns = ['Date', 'ES', 'Document', 'Tiers', 'Reference', 'Quantite', 'Unite',
                      'PU_HT', 'Solde', 'CUMP', 'Depot', 'Categorie', 'Sous_Cat', 'Marque'][:ncols]
    else:
        df.columns = ['Date', 'ES', 'Document', 'Tiers', 'Reference', 'Quantite', 'Unite',
                      'PU_HT', 'Solde', 'Depot', 'Categorie', 'Sous_Cat', 'Marque'][:ncols]

    df['Qty'] = parse_qty_series(df['Quantite'])
    df['Date'] = parse_dates_series(df['Date'])
    df = df[df['ES'].isin(['S', 'E', 'ME', 'TE'])]
    df = df.drop_duplicates(subset=['Date', 'ES', 'Document', 'Reference', 'Quantite'])
    return df


def charger_mouv_upload(file_bytes):
    df = pd.read_excel(file_bytes)
    return normaliser_mouv(df)


@st.cache_resource(show_spinner=False)
def charger_base_historique():
    """Base 2020-2025 livrée avec l'app — lue une seule fois, gardée en mémoire serveur."""
    return pd.read_pickle(BASE_HISTORIQUE)


def _stock_moyen_pondere(events_df, date_debut, date_fin, stock_actuel):
    """
    Reconstruit le STOCK MOYEN réel (pondéré dans le temps) sur [date_debut, date_fin],
    par référence, à partir :
      - du Stock actuel (connu avec certitude à date_fin, issu de l'export ERP du jour) ;
      - de l'historique des mouvements (entrées positives, sorties négatives) survenus
        dans la fenêtre, avec leur date exacte.

    Logique : en partant du niveau connu à date_fin, on "remonte" mouvement par
    mouvement pour reconstituer le niveau de stock à chaque instant du passé, puis on
    calcule la moyenne pondérée par la durée de chaque palier (et non une simple
    moyenne arithmétique des mouvements, qui ignorerait le temps réellement passé à
    chaque niveau de stock).

    Hypothèse (à connaître) : cette reconstruction suppose que les mouvements ES
    ('S','E','ME','TE') couvrent bien toutes les variations de stock sur la période
    (pas d'ajustement d'inventaire hors mouvement, pas de perte/casse non enregistrée).
    Les niveaux reconstruits sont plafonnés à 0 (un stock ne peut pas être négatif) —
    si ce plafonnage se déclenche souvent, c'est le signe d'un écart entre la date de
    l'export stock et la couverture réelle des mouvements.
    """
    total_duration = (date_fin - date_debut).total_seconds()
    resultat = stock_actuel.astype(float).copy()
    if total_duration <= 0 or events_df.empty:
        return resultat

    for ref, g in events_df.sort_values('Date').groupby('Reference'):
        if ref not in resultat.index:
            continue
        s_now = stock_actuel.get(ref, 0.0)
        niveau = s_now - g['Delta'].sum()  # niveau au début de la fenêtre
        t_prec = date_debut
        somme_ponderee = 0.0
        for d, delta in zip(g['Date'], g['Delta']):
            duree = (d - t_prec).total_seconds()
            if duree > 0:
                somme_ponderee += max(niveau, 0.0) * duree
            niveau += delta
            t_prec = d
        duree = (date_fin - t_prec).total_seconds()
        if duree > 0:
            somme_ponderee += max(niveau, 0.0) * duree
        resultat.loc[ref] = somme_ponderee / total_duration

    return resultat


@st.cache_data(show_spinner=False)
def calculer_indicateurs(df_mv, df_st_bytes):
    df_st = pd.read_excel(df_st_bytes)
    # L'unité fiable est celle d'ACHAT ("Unité P.") — l'unité de VENTE ("Unité V.") contient
    # des erreurs de saisie ponctuelles (ex: une référence en M2 étiquetée "P" à la vente).
    _unite_col = next((c for c in df_st.columns if str(c).strip().upper() == 'UNITÉ P.'), None)
    if _unite_col is None:
        _unite_col = next(
            (c for c in df_st.columns if str(c).strip().upper().startswith('UNIT') and 'P' in str(c).upper()),
            None)
    if _unite_col is None:
        _unite_col = next(
            (c for c in df_st.columns if str(c).strip().upper().startswith('UNIT') and 'V' in str(c).upper()),
            None)
    if _unite_col and _unite_col != 'Unité':
        df_st.rename(columns={_unite_col: 'Unité'}, inplace=True)
    elif _unite_col is None:
        df_st['Unité'] = 'M3'
    if 'Désignation' not in df_st.columns:
        df_st['Désignation'] = ''
    df_st['Qty_s'] = parse_qty_series(df_st['Quantité'])
    df_st_c = df_st.groupby('Référence').agg(
        Stock=('Qty_s', 'sum'), Cat=('Catégorie', 'first'), Unite=('Unité', 'first'),
        Designation=('Désignation', 'first')).reset_index()
    df_st_c = df_st_c[df_st_c['Cat'].notna()]
    # Par défaut BOIS BLANC/BOIS ROUGE sont en M3, mais certaines références (ex: SCHAAL)
    # sont réellement vendues en ML — on ne force M3 que si l'unité réelle n'est ni M3 ni ML,
    # pour ne pas écraser une unité déjà correcte (ex: HOLMEN_25X100_MB-SCHAAL en ML).
    _unite_reconnue = df_st_c['Unite'].astype(str).str.upper().isin(['M3', 'M³', 'ML', 'M/L', 'M.L'])
    df_st_c.loc[df_st_c['Cat'].isin(['BOIS BLANC', 'BOIS ROUGE']) & ~_unite_reconnue, 'Unite'] = 'M3'

    # ── Normalisation des références de mouvements historiques ──
    # D'anciens exports (2020-2025) enregistrent certaines lignes BOIS ROUGE/BOIS BLANC avec le
    # préfixe de catégorie collé à la référence (ex: "BOIS ROUGE KAJA_38X150_US" au lieu de
    # "KAJA_38X150_US"). Comme la fusion avec le stock se fait par égalité stricte de référence,
    # ces mouvements restaient invisibles (article classé "Sans mouvement" à tort). On ne
    # dé-préfixe que si la référence nue existe réellement dans le stock actuel, pour ne jamais
    # créer de faux rapprochement.
    _valid_refs = set(df_st_c['Référence'].astype(str))
    _ref_mv = df_mv['Reference'].astype(str)
    _mask_deja_valide = _ref_mv.isin(_valid_refs)
    for _prefix in ('BOIS ROUGE ', 'BOIS BLANC '):
        _mask_prefixe = ~_mask_deja_valide & _ref_mv.str.upper().str.startswith(_prefix)
        if _mask_prefixe.any():
            _stripped = _ref_mv[_mask_prefixe].str.slice(len(_prefix)).str.strip()
            _mask_recupere = _stripped.isin(_valid_refs)
            df_mv.loc[_stripped[_mask_recupere].index, 'Reference'] = _stripped[_mask_recupere]
            _ref_mv = df_mv['Reference'].astype(str)
            _mask_deja_valide = _ref_mv.isin(_valid_refs)

    date_max = df_mv['Date'].max()
    date_12m_debut = date_max - pd.DateOffset(months=12)

    sorties = df_mv[df_mv['ES'] == 'S']
    entrees = df_mv[df_mv['ES'].isin(['E', 'ME', 'TE'])]

    mv_s_all = sorties.groupby('Reference').agg(
        Total_Sorti=('Qty', 'sum'), Nb_Trans=('Qty', 'count'), Dern_Sortie=('Date', 'max')).reset_index()
    mv_e_all = entrees.groupby('Reference').agg(Dern_Entree=('Date', 'max')).reset_index()

    s_12m_df = sorties[(sorties['Date'] >= date_12m_debut) & (sorties['Date'] <= date_max)]
    e_12m_df = entrees[(entrees['Date'] >= date_12m_debut) & (entrees['Date'] <= date_max)]
    mv_s_12m = s_12m_df.groupby('Reference').agg(S_12M=('Qty', 'sum'), T_12M=('Qty', 'count')).reset_index()
    mv_e_12m = e_12m_df.groupby('Reference')['Qty'].sum().rename('A_12M').reset_index()

    # ── Fenêtre 4 mois (tendance récente) ──
    date_4m_debut = date_max - pd.DateOffset(months=4)
    s_4m_df = sorties[(sorties['Date'] >= date_4m_debut) & (sorties['Date'] <= date_max)]
    e_4m_df = entrees[(entrees['Date'] >= date_4m_debut) & (entrees['Date'] <= date_max)]
    mv_s_4m = s_4m_df.groupby('Reference')['Qty'].sum().rename('S_4M').reset_index()

    # ── Historique des mouvements signés (entrée = +, sortie = -), pour reconstituer
    #    le stock moyen réel sur chaque fenêtre à partir du stock actuel connu ──
    events_12m = pd.concat([
        s_12m_df[['Reference', 'Date', 'Qty']].assign(Delta=lambda d: -d['Qty']),
        e_12m_df[['Reference', 'Date', 'Qty']].assign(Delta=lambda d: d['Qty']),
    ], ignore_index=True) if (len(s_12m_df) or len(e_12m_df)) else pd.DataFrame(columns=['Reference', 'Date', 'Qty', 'Delta'])
    events_4m = pd.concat([
        s_4m_df[['Reference', 'Date', 'Qty']].assign(Delta=lambda d: -d['Qty']),
        e_4m_df[['Reference', 'Date', 'Qty']].assign(Delta=lambda d: d['Qty']),
    ], ignore_index=True) if (len(s_4m_df) or len(e_4m_df)) else pd.DataFrame(columns=['Reference', 'Date', 'Qty', 'Delta'])

    # ── Historique par année (S_2020, S_2021 ... / A_2020, A_2021 ...) ──
    sorties_y = sorties.assign(Annee=sorties['Date'].dt.year)
    entrees_y = entrees.assign(Annee=entrees['Date'].dt.year)
    piv_s = sorties_y.groupby(['Reference', 'Annee'])['Qty'].sum().unstack(fill_value=0)
    piv_t = sorties_y.groupby(['Reference', 'Annee'])['Qty'].count().unstack(fill_value=0)
    piv_a = entrees_y.groupby(['Reference', 'Annee'])['Qty'].sum().unstack(fill_value=0)
    piv_s.columns = [f'S_{int(y)}' for y in piv_s.columns]
    piv_t.columns = [f'T_{int(y)}' for y in piv_t.columns]
    piv_a.columns = [f'A_{int(y)}' for y in piv_a.columns]
    cols_annuelles = sorted(set(piv_s.columns) | set(piv_a.columns) | set(piv_t.columns))

    sm = df_st_c.copy()
    sm = sm.merge(mv_s_all, left_on='Référence', right_on='Reference', how='left').drop(columns='Reference', errors='ignore')
    sm = sm.merge(mv_e_all, left_on='Référence', right_on='Reference', how='left').drop(columns='Reference', errors='ignore')
    sm = sm.merge(mv_s_12m, left_on='Référence', right_on='Reference', how='left').drop(columns='Reference', errors='ignore')
    sm = sm.merge(mv_e_12m, left_on='Référence', right_on='Reference', how='left').drop(columns='Reference', errors='ignore')
    sm = sm.merge(mv_s_4m, left_on='Référence', right_on='Reference', how='left').drop(columns='Reference', errors='ignore')
    sm = sm.merge(piv_s, left_on='Référence', right_index=True, how='left')
    sm = sm.merge(piv_t, left_on='Référence', right_index=True, how='left')
    sm = sm.merge(piv_a, left_on='Référence', right_index=True, how='left')
    for c in ['Total_Sorti', 'Nb_Trans', 'S_12M', 'T_12M', 'A_12M', 'S_4M'] + cols_annuelles:
        sm[c] = sm[c].fillna(0)

    stock_ok = sm['Stock'] > SEUIL

    # ── Stock moyen réel 12M / 4M (reconstruit depuis le stock actuel + l'historique
    #    des mouvements datés — voir _stock_moyen_pondere) ──
    stock_actuel_series = sm.set_index('Référence')['Stock']
    stock_moyen_12m = _stock_moyen_pondere(events_12m, date_12m_debut, date_max, stock_actuel_series)
    stock_moyen_4m = _stock_moyen_pondere(events_4m, date_4m_debut, date_max, stock_actuel_series)
    sm['Stock_moyen_12M'] = sm['Référence'].map(stock_moyen_12m).fillna(sm['Stock']).round(3)
    sm['Stock_moyen_4M'] = sm['Référence'].map(stock_moyen_4m).fillna(sm['Stock']).round(3)

    # ── Moyennes mensuelles ──
    sm['Moy_Mois_12M'] = (sm['S_12M'] / 12.0).round(3)
    sm['Moy_Mois_4M'] = (sm['S_4M'] / 4.0).round(3)

    # ── Rotation actuelle = Sorties 12M ÷ Stock ACTUEL (même formule que le KPI
    #    Dashboard, au niveau article). Sert de repère "vue instantanée" à côté de la
    #    Rotation 12M historique — les deux sont affichées séparément, jamais fondues. ──
    sm['Rotation_Actuelle'] = 0.0
    m = stock_ok & (sm['S_12M'] > 0)
    sm.loc[m, 'Rotation_Actuelle'] = (sm.loc[m, 'S_12M'] / sm.loc[m, 'Stock']).round(2)

    # ── Rotation 12M et 4M HISTORIQUES = Sorties ÷ STOCK MOYEN de la période (pas le
    #    stock actuel), ce qui évite l'explosion du ratio quand le stock actuel est
    #    quasi nul. Indicateur analytique — ne pilote jamais seul la Classification. ──
    sm['Rotation_12M'] = 0.0
    m = (sm['Stock_moyen_12M'] > SEUIL) & (sm['S_12M'] > 0)
    sm.loc[m, 'Rotation_12M'] = (sm.loc[m, 'S_12M'] / sm.loc[m, 'Stock_moyen_12M']).round(2)

    sm['Rotation_4M'] = 0.0
    m = (sm['Stock_moyen_4M'] > SEUIL) & (sm['S_4M'] > 0)
    sm.loc[m, 'Rotation_4M'] = (sm.loc[m, 'S_4M'] / sm.loc[m, 'Stock_moyen_4M']).round(2)

    # ── Tendance de la demande : Moy/Mois 4M vs Moy/Mois 12M ──
    sm['Tendance_Ratio'] = float('nan')
    m = sm['Moy_Mois_12M'] > 0
    sm.loc[m, 'Tendance_Ratio'] = sm.loc[m, 'Moy_Mois_4M'] / sm.loc[m, 'Moy_Mois_12M']
    sm['Tendance_Pct'] = (sm['Tendance_Ratio'] - 1) * 100

    def _label_tendance(row):
        ratio, pct = row['Tendance_Ratio'], row['Tendance_Pct']
        if pd.isna(ratio):
            return '📈 Hausse (nouvelle activité)' if row['Moy_Mois_4M'] > 0 else '⚪ Aucune demande'
        if ratio > 1.10:
            return f'📈 Hausse de la demande ({pct:+.0f}%)'
        if ratio < 0.90:
            return f'📉 Baisse de la demande ({pct:+.0f}%)'
        return f'➡️ Demande stable ({pct:+.0f}%)'
    sm['Tendance_Label'] = sm.apply(_label_tendance, axis=1)

    # ── Couverture = Stock actuel ÷ Moy/Mois 4M (mesure la tenue face à la demande
    #    RÉCENTE, pas historique). NaN si aucune sortie sur 4M : la notion de "nombre
    #    de mois de couverture" n'a pas de sens si rien ne sort — géré via la
    #    Classification (Dormant / Sans mouvement) plutôt qu'inventé à 0 ou l'infini ──
    sm['Couverture'] = float('nan')
    sm.loc[~stock_ok, 'Couverture'] = 0.0
    m = stock_ok & (sm['Moy_Mois_4M'] > 0)
    sm.loc[m, 'Couverture'] = (sm.loc[m, 'Stock'] / sm.loc[m, 'Moy_Mois_4M']).round(1)

    sm['Taux_Immob'] = 0.0
    m = sm['Couverture'] > 0
    sm.loc[m, 'Taux_Immob'] = (sm.loc[m, 'Couverture'] / 12 * 100).clip(upper=100.0).round(1)
    sm.loc[(~m) & stock_ok, 'Taux_Immob'] = 100.0

    # ── Classification : priorité à la Couverture / au niveau de stock, plus jamais
    #    une Rotation élevée seule. Un article sans sortie récente (4M) n'est jamais
    #    classé comme "bon niveau" au seul motif d'une grande couverture. ──
    cov = sm['Couverture']
    conditions_rupture = ~stock_ok
    conditions_sans_mouvement = stock_ok & (sm['S_4M'] == 0) & (sm['Total_Sorti'] == 0)
    conditions_dormant = stock_ok & (sm['S_4M'] == 0) & (sm['Total_Sorti'] > 0)
    m_actif = stock_ok & (sm['S_4M'] > 0)

    sm['Class'] = 'Normal'
    sm.loc[m_actif & (cov < 1), 'Class'] = 'Critique'
    sm.loc[m_actif & (cov >= 1) & (cov < 3), 'Class'] = 'Stock faible'
    sm.loc[m_actif & (cov >= 3) & (cov < 6), 'Class'] = 'Normal'
    sm.loc[m_actif & (cov >= 6) & (cov < 12), 'Class'] = 'Bon niveau'
    sm.loc[m_actif & (cov >= 12) & (cov < 18), 'Class'] = 'Surstock'
    sm.loc[m_actif & (cov >= 18), 'Class'] = 'Surstock important'
    sm.loc[conditions_dormant, 'Class'] = 'Dormant'
    sm.loc[conditions_sans_mouvement, 'Class'] = 'Sans mouvement'
    sm.loc[conditions_rupture, 'Class'] = 'Rupture'

    return sm, date_max, date_12m_debut, cols_annuelles




def format_nombre_fr(value, decimals=2):
    """Formatage sobre des nombres pour l'affichage des indicateurs."""
    if pd.isna(value):
        return "—"
    return f"{value:,.{decimals}f}".replace(',', ' ')


def format_unite_stock(unite):
    """Libellé court des unités affichées dans les KPI, sans modifier les calculs."""
    u = str(unite).strip().upper()
    return {
        'M3': 'm³', 'M³': 'm³', 'M2': 'm²', 'M²': 'm²',
        'P': 'P', 'PC': 'P', 'PCS': 'P', 'PIECE': 'P', 'PIÈCE': 'P',
        'ML': 'ML', 'M/L': 'ML', 'M.L': 'ML',
    }.get(u, u or 'Unité')

def evolution_mensuelle(df_mv, references=None):
    d = df_mv[df_mv['ES'] == 'S']
    if references is not None:
        d = d[d['Reference'].isin(references)]
    d = d.copy()
    d['Mois'] = d['Date'].dt.to_period('M').dt.to_timestamp()
    return d.groupby('Mois', as_index=False)['Qty'].sum()



def make_pdf_bytes(title, df, date_max):
    """Génère un PDF texte simple sans dépendance externe."""
    lines = [f"WOODMAT - {title}", f"Analyse du {date_max.strftime('%d/%m/%Y')}", ""]
    lines.extend([" | ".join(map(str, row[:8])) for row in df.head(80).itertuples(index=False, name=None)])
    content = "BT /F1 10 Tf 40 800 Td "
    escaped = []
    for line in lines:
        safe = str(line).replace('\\', '\\\\').replace('(', '\\(').replace(')', '\\)')[:120]
        escaped.append(f"({safe}) Tj 0 -14 Td")
    stream = (content + " ".join(escaped) + " ET").encode('latin-1', errors='replace')
    objects = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj",
        b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj",
        b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj",
        b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj",
        b"5 0 obj << /Length " + str(len(stream)).encode() + b" >> stream\n" + stream + b"\nendstream endobj",
    ]
    pdf = b"%PDF-1.4\n"
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf)); pdf += obj + b"\n"
    xref = len(pdf)
    pdf += f"xref\n0 {len(objects)+1}\n0000000000 65535 f \n".encode()
    for off in offsets[1:]:
        pdf += f"{off:010d} 00000 n \n".encode()
    pdf += f"trailer << /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF".encode()
    return pdf


def add_export_buttons(df, basename, sheet_name, date_max):
    c1, c2 = st.columns([1, 1])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    c1.download_button("⬇️ Exporter Excel", buf.getvalue(),
                       file_name=f"{basename}_{date_max.strftime('%d_%m_%Y')}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    c2.download_button("⬇️ Exporter PDF", make_pdf_bytes(sheet_name, df, date_max),
                       file_name=f"{basename}_{date_max.strftime('%d_%m_%Y')}.pdf",
                       mime="application/pdf")


def replenishment_action(couverture):
    if pd.isna(couverture):
        return "⚪ Sans sortie récente (4M) — à vérifier"
    if couverture < 1:
        return "🔴 Commander immédiatement"
    if couverture <= 2:
        return "🟠 Commander prochainement"
    return "🟢 Rien à faire"


def charger_parametres_stock():
    """Charge parametres_stock.json (paramètres par catégorie). {} si absent/invalide."""
    if os.path.exists(PARAMS_FILE):
        try:
            with open(PARAMS_FILE, 'r', encoding='utf-8') as fp:
                return json.load(fp)
        except Exception:
            return {}
    return {}


def get_params_categorie(cat, params):
    """Paramètres effectifs d'une catégorie : valeurs saisies dans Paramètres,
    sinon repli sur DEFAULT_PARAMS_CATEGORIE (jamais déduites des données)."""
    p = (params or {}).get(cat, {})
    return {
        'delai_appro_mois': float(p.get('lead_time_mois', DEFAULT_PARAMS_CATEGORIE['lead_time_mois'])),
        'seuil_rupture_mois': float(p.get('seuil_rupture', DEFAULT_PARAMS_CATEGORIE['seuil_rupture'])),
        'stock_securite_mois': float(p.get('stock_securite', DEFAULT_PARAMS_CATEGORIE['stock_securite'])),
        'stock_cible_mois': float(p.get('stock_cible', DEFAULT_PARAMS_CATEGORIE['stock_cible'])),
    }


def calculer_moteur_reappro(df, params=None):
    """
    Moteur de réapprovisionnement — applique les 4 paramètres définis PAR CATÉGORIE
    (Délai d'approvisionnement, Seuil de rupture, Stock de sécurité, Stock cible) à
    chaque référence de `df`. Les 4 restent toujours distincts (jamais fondus l'un
    dans l'autre).

    Consommation mensuelle moyenne = Moy_Mois_4M, déjà calculée dans
    calculer_indicateurs() à partir des sorties réelles des 4 derniers mois — jamais
    inventée. Si elle est nulle (aucune sortie récente), la recommandation est
    marquée explicitement non fiable plutôt que masquée ou forcée à une valeur.

    Ne modifie aucune donnée source de mouvements : ajoute uniquement des colonnes
    calculées à une COPIE du DataFrame fourni.
    """
    if params is None:
        params = charger_parametres_stock()
    d = df.copy()

    par_cat = {cat: get_params_categorie(cat, params) for cat in d['Cat'].dropna().unique()}
    d['Delai_Appro_Mois'] = d['Cat'].map(lambda c: par_cat.get(c, get_params_categorie(c, params))['delai_appro_mois'])
    d['Seuil_Rupture_Mois'] = d['Cat'].map(lambda c: par_cat.get(c, get_params_categorie(c, params))['seuil_rupture_mois'])
    d['Stock_Securite_Mois'] = d['Cat'].map(lambda c: par_cat.get(c, get_params_categorie(c, params))['stock_securite_mois'])
    d['Stock_Cible_Mois'] = d['Cat'].map(lambda c: par_cat.get(c, get_params_categorie(c, params))['stock_cible_mois'])

    conso = d['Moy_Mois_4M'] if 'Moy_Mois_4M' in d.columns else pd.Series(0.0, index=d.index)
    d['Conso_Mensuelle_Moyenne'] = conso.fillna(0.0)
    d['Demande_Recente'] = d['Conso_Mensuelle_Moyenne'] > 0

    # Couverture (mois) = Stock actuel ÷ Consommation mensuelle moyenne réelle (4M)
    d['Couverture_Reappro'] = float('nan')
    m = d['Demande_Recente']
    d.loc[m, 'Couverture_Reappro'] = (d.loc[m, 'Stock'] / d.loc[m, 'Conso_Mensuelle_Moyenne']).round(2)

    def _risque(row):
        if not row['Demande_Recente']:
            return '⚪ Sans demande récente — recommandation non fiable'
        cov = row['Couverture_Reappro']
        if cov <= row['Seuil_Rupture_Mois']:
            return '🔴 Risque de rupture'
        if cov <= row['Delai_Appro_Mois']:
            return '⚠️ Commander maintenant (avant réception)'
        return '🟢 Couverture suffisante'
    d['Risque_Reappro'] = d.apply(_risque, axis=1)

    # Quantité recommandée = max(0, Stock cible × Conso mensuelle moyenne − Stock actuel)
    d['Qte_Recommandee'] = 0.0
    d.loc[m, 'Qte_Recommandee'] = (
        d.loc[m, 'Stock_Cible_Mois'] * d.loc[m, 'Conso_Mensuelle_Moyenne'] - d.loc[m, 'Stock']
    ).clip(lower=0).round(2)

    # Stock de sécurité : signal distinct (protection contre la variabilité), jamais
    # fondu dans la quantité recommandée ci-dessus.
    d['Sous_Stock_Securite'] = False
    d.loc[m, 'Sous_Stock_Securite'] = d.loc[m, 'Stock'] < (d.loc[m, 'Stock_Securite_Mois'] * d.loc[m, 'Conso_Mensuelle_Moyenne'])

    d['Fiabilite'] = d['Demande_Recente'].map({True: 'Fiable (4M réels)', False: 'Non fiable — sans demande récente'})
    return d


# ============================================================
# FEUILLE STOCK BOIS ROUGE (Qualité × Fournisseur × Dimension)
# ENSO et STORA ENSO sont le même fournisseur → fusionnés en une seule colonne
# ============================================================

import re as _re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BR_QUAL = ['US', 'V', 'VI', 'VII', 'SCHAAL']
BR_MAIN = ['UPM', 'STORA ENSO', 'KAJA', 'WISA', 'KAUKAS', 'HASA', 'KEITELE', 'JULA', 'SEIKKU']
BR_QCOL = {
    'US': ('1F4E79', 'AED6F1'), 'V': ('196F3D', 'A9DFBF'), 'VI': ('6C3483', 'D7BDE2'),
    'VII': ('B7410E', 'F5CBA7'), 'SCHAAL': ('17594A', 'A2D9CE'),
}
BR_TITLE, BR_TOT, BR_SUB, BR_GRAND = '5C2D0A', '8B4513', 'D5B49A', 'A0522D'
BR_DIM1, BR_DIM2 = 'FDF0E8', 'FFFFFF'


def _br_extract_qual(ref):
    for q in ['SCHAAL', 'VII', 'VI', 'V', 'US']:
        if _re.search(r'[\s_]' + q + r'$', str(ref)):
            return q
    return None


def _br_extract_four(ref):
    ref2 = _re.sub(r'^BOIS\s+ROUGE\s+', '', str(ref).upper().strip())
    # ENSO et STORA ENSO fusionnés — même fournisseur
    for f, canon in [('STORA ENSO', 'STORA ENSO'), ('STORA', 'STORA ENSO'), ('ENSO', 'STORA ENSO'),
                      ('KAUKAS', 'KAUKAS'), ('WISA', 'WISA'), ('KAJA', 'KAJA'), ('UPM', 'UPM'),
                      ('JULA', 'JULA'), ('HASA', 'HASA'), ('KEITELE', 'KEITELE'), ('SEIKKU', 'SEIKKU')]:
        if ref2.startswith(f):
            return canon
    return None


def _br_extract_dim(ref, texte5=None):
    ref2 = _re.sub(r'\s*[Xx]\s*', 'X', str(ref).upper())
    m = _re.search(r'(\d{2,3}X\d{2,3})', ref2)
    if m:
        return m.group(1)
    if texte5 is not None and pd.notna(texte5):
        return _re.sub(r'\s*[Xx]\s*', 'X', str(texte5).strip().upper())
    return 'INCONNU'


def _br_dim_key(d):
    p = _re.findall(r'\d+', d)
    return [int(x) for x in p] if p else [0]


def generer_excel_bois_rouge(df_st_raw, date_max):
    """Reconstruit la feuille Stock Bois Rouge (style identique à l'ancien outil desktop),
    avec ENSO et STORA ENSO fusionnés en un seul fournisseur."""
    # BOIS ROUGE + BOIS BLANC SUEDE (même famille ENSO/STORA ENSO) réunis dans le même tableau
    df_br = df_st_raw[df_st_raw['Catégorie'].isin(['BOIS ROUGE', 'BOIS BLANC SUEDE'])].copy()
    df_br['Quantité'] = parse_qty_series(df_br['Quantité'])
    df_br['_qual'] = df_br['Référence'].apply(_br_extract_qual)
    df_br['_four'] = df_br['Référence'].apply(_br_extract_four)
    if 'Texte 5' in df_br.columns:
        df_br['_dim'] = df_br.apply(lambda r: _br_extract_dim(r['Référence'], r.get('Texte 5')), axis=1)
    else:
        df_br['_dim'] = df_br['Référence'].apply(lambda r: _br_extract_dim(r))

    non_reconnus = df_br[(df_br['Quantité'] > 0) & df_br['_qual'].notna() & df_br['_four'].isna()]
    df_br = df_br[(df_br['Quantité'] > 0) & df_br['_qual'].notna() & df_br['_four'].notna()].copy()

    if df_br.empty:
        return None, non_reconnus

    br_qf = {}
    for q in BR_QUAL:
        present = df_br[df_br['_qual'] == q]['_four'].unique()
        br_qf[q] = [f for f in BR_MAIN if f in present] + sorted([f for f in present if f not in BR_MAIN])

    br_dims = sorted(df_br['_dim'].unique(), key=_br_dim_key)
    br_pivot = df_br.pivot_table(index='_dim', columns=['_qual', '_four'], values='Quantité',
                                  aggfunc='sum', fill_value=0).reindex(br_dims)

    def gv(dim, q, fr):
        try:
            k = (q, fr)
            if k in br_pivot.columns:
                v = br_pivot.loc[dim, k]
                return float(v) if pd.notna(v) else 0.0
        except Exception:
            pass
        return 0.0

    br_flat = []
    for q in BR_QUAL:
        for fr in br_qf[q]:
            br_flat.append((q, fr, 'data'))
        br_flat.append((q, 'TOTAL', 'subtotal'))
    br_flat.append(('', 'TOTAL', 'grandtotal'))

    def mf(h):
        return PatternFill("solid", fgColor=h)
    thin = Side(style='thin', color='BBBBBB')
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Stock BOIS ROUGE")
    ws.sheet_view.showGridLines = False
    n_cols = len(br_flat)
    dt = date_max.strftime('%d/%m/%Y')

    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + n_cols)
    c = ws.cell(1, 1, f"WOODMAT  —  STOCK BOIS ROUGE + BOIS BLANC SUEDE (ENSO)  —  Quantités en M³  |  {dt}")
    c.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    c.fill = mf(BR_TITLE)
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 20
    ws.cell(2, 1, '').fill = mf(BR_TITLE)
    cc = 2
    for q in BR_QUAL:
        n = len(br_qf[q]) + 1
        ws.merge_cells(start_row=2, start_column=cc, end_row=2, end_column=cc + n - 1)
        c = ws.cell(2, cc, q)
        c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
        c.fill = mf(BR_QCOL[q][0])
        c.alignment = Alignment(horizontal='center', vertical='center')
        cc += n
    c = ws.cell(2, cc, 'TOTAL')
    c.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    c.fill = mf(BR_TOT)
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[3].height = 20
    c = ws.cell(3, 1, 'DIMENSION')
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = mf(BR_TITLE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd
    ws.column_dimensions['A'].width = 13

    for ci, (q, fr, ctype) in enumerate(br_flat):
        col = 2 + ci
        bg = BR_TOT if ctype == 'grandtotal' else (BR_QCOL[q][0] if ctype == 'subtotal' else BR_QCOL[q][1])
        fg = 'FFFFFF' if ctype != 'data' else '1A1A1A'
        c = ws.cell(3, col, fr)
        c.font = Font(name='Arial', bold=(ctype != 'data'), size=8, color=fg)
        c.fill = mf(bg)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = brd
        ws.column_dimensions[get_column_letter(col)].width = 8 if ctype == 'data' else 10

    for ri, dim in enumerate(br_dims):
        row = ri + 4
        ws.row_dimensions[row].height = 15
        dim_bg = BR_DIM1 if ri % 2 == 0 else BR_DIM2
        c = ws.cell(row, 1, dim)
        c.font = Font(name='Arial', size=9, color='3B1500')
        c.fill = mf(dim_bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = brd

        row_grand = 0.0
        for ci, (q, fr, ctype) in enumerate(br_flat):
            col = 2 + ci
            if ctype == 'data':
                val = gv(dim, q, fr)
            elif ctype == 'subtotal':
                val = sum(gv(dim, q, ff) for ff in br_qf[q])
                row_grand += val
            else:
                val = row_grand
            c = ws.cell(row, col)
            if val < 0.0005:
                c.value = None
                c.fill = mf('F2F2F2' if ctype == 'data' else ('E0D0C4' if ctype == 'subtotal' else 'E8CEB8'))
            else:
                c.value = round(val, 3)
                c.number_format = '#,##0.000'
                c.fill = mf(BR_GRAND if ctype == 'grandtotal' else (BR_SUB if ctype == 'subtotal' else dim_bg))
            c.font = Font(name='Arial', bold=(ctype != 'data'), size=9,
                          color=('FFFFFF' if ctype == 'grandtotal' else ('5C2D0A' if ctype == 'subtotal' else '3B1500')))
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = brd

    tot_row = len(br_dims) + 4
    ws.row_dimensions[tot_row].height = 18
    c = ws.cell(tot_row, 1, 'TOTAL')
    c.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    c.fill = mf(BR_TOT)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd

    grand_col_total = 0.0
    for ci, (q, fr, ctype) in enumerate(br_flat):
        col = 2 + ci
        if ctype == 'data':
            col_sum = sum(gv(dim, q, fr) for dim in br_dims)
        elif ctype == 'subtotal':
            col_sum = sum(sum(gv(dim, q, ff) for ff in br_qf[q]) for dim in br_dims)
            grand_col_total += col_sum
        else:
            col_sum = grand_col_total
        c = ws.cell(tot_row, col)
        c.value = round(col_sum, 3) if col_sum > 0.0005 else None
        c.number_format = '#,##0.000'
        c.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        c.fill = mf(BR_TOT)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = brd

    ws.freeze_panes = 'B4'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), non_reconnus


# ============================================================
# INTERFACE
# ============================================================

if not check_login():
    st.stop()

if not os.path.exists(BASE_HISTORIQUE):
    st.error("⚠️ La base historique (base_mouvements.pkl) n'est pas présente dans le déploiement. "
             "Elle doit être livrée avec l'application — contactez l'administrateur.")
    st.stop()

inject_global_styles()

MENU_ITEMS = [
    "🏠 Dashboard", "📦 Rotation du stock", "📦 Réapprovisionnement", "📈 Analyses", "⚙️ Paramètres", "⚠️ Alertes",
    "📄 Rapports", "📚 Historique", "😴 Stock dormant", "🪵 Stock Bois Rouge"
]
if user_can_manage_users():
    MENU_ITEMS.append("👥 Gestion des utilisateurs")

with st.sidebar:
    st.markdown("### 🪵 WOODMAT")
    st.caption(f"Connecté : {get_current_user().get('name', '')}")
    st.divider()
    page = st.radio("Navigation", MENU_ITEMS, label_visibility="collapsed", key="nav_menu")
    st.divider()
    st.caption("La base historique 2020–2025 est intégrée à l'application — rien à importer.")

    f_mouv = None
    f_stock = None
    lancer = False
    # Show uploaders only on pages that need data imports/analysis generation.
    if page in ["🏠 Dashboard", "📦 Rotation du stock", "📦 Réapprovisionnement", "📈 Analyses", "📚 Historique", "😴 Stock dormant", "🪵 Stock Bois Rouge"]:
        st.markdown("**1. Mouvements de l'année en cours** _(optionnel)_")
        f_mouv = st.file_uploader("Export ERP mouvements (ex : 2026)", type=["xlsx", "xls"], key="mouv")

        st.markdown("**2. Stock actuel** _(obligatoire, export du jour)_")
        f_stock = st.file_uploader("Export ERP stock actuel", type=["xlsx", "xls"], key="stock")

        lancer = st.button("🔄 Générer l'analyse", type="primary", use_container_width=True)

current_user = get_current_user()
user = current_user.get('name', current_user.get('email', ''))
role = current_user.get('role', '—')
today = pd.Timestamp.now().strftime('%d/%m/%Y')
st.markdown(
    f"""
    <div class='woodmat-header'>
        <div style='display:flex;align-items:center;gap:0.75rem'>
            <div class='woodmat-logo' style='width:42px;height:42px;border-radius:13px;font-size:1.2rem;margin:0'>W</div>
            <div><div class='woodmat-header-title'>WOODMAT — Rotation du stock</div><div class='woodmat-muted'>Application métier</div></div>
        </div>
        <div class='woodmat-header-meta'>📅 {today}<br>👤 {user}<br>🔐 {role}</div>
    </div>
    """,
    unsafe_allow_html=True)
if st.button("Déconnexion", key="logout_header"):
    st.session_state.clear()
    st.rerun()

if "sm" not in st.session_state:
    st.session_state["sm"] = None

if lancer:
    if f_stock is None:
        st.error("Choisissez le fichier stock actuel.")
    else:
        with st.spinner("Calcul des indicateurs de rotation..."):
            df_base = charger_base_historique()
            if f_mouv is not None:
                df_recent = charger_mouv_upload(f_mouv)
                df_mv = pd.concat([df_base, df_recent], ignore_index=True)
                df_mv = df_mv.drop_duplicates(subset=['Date', 'ES', 'Document', 'Reference', 'Quantite'])
                # Fusion + sauvegarde automatique de la base pour les prochaines analyses
                df_mv.to_pickle(BASE_HISTORIQUE)
                charger_base_historique.clear()
            else:
                df_mv = df_base
            sm, date_max, date_12m_debut, cols_annuelles = calculer_indicateurs(df_mv, f_stock)
            f_stock.seek(0)
            df_st_raw = pd.read_excel(f_stock)
            st.session_state["sm"] = sm
            st.session_state["df_mv"] = df_mv
            st.session_state["date_max"] = date_max
            st.session_state["date_12m_debut"] = date_12m_debut
            st.session_state["cols_annuelles"] = cols_annuelles
            st.session_state["df_st_raw"] = df_st_raw
            st.session_state.setdefault("analysis_history", []).append({
                "Date": pd.Timestamp.now().strftime('%d/%m/%Y %H:%M'),
                "Utilisateur": get_current_user().get("name", ""),
                "Catégorie": "Toutes",
                "Nombre d'articles": len(sm),
                "Durée d'analyse": "—",
            })
        if f_mouv is not None:
            st.success("Mouvements fusionnés et base historique mise à jour sur le serveur ✅")

sm = st.session_state.get("sm")

if page == "👥 Gestion des utilisateurs":
    render_user_management()
    st.stop()

if page == "⚙️ Paramètres":
    st.markdown("<h2 class='woodmat-page-title'>⚙️ Paramètres</h2>", unsafe_allow_html=True)
    params_file = PARAMS_FILE

    def load_params():
        if os.path.exists(params_file):
            try:
                with open(params_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception:
                return {}
        return {}

    params = load_params()

    # categories detected from last analysis or from raw stock
    df_st_raw = st.session_state.get('df_st_raw')
    if st.session_state.get('sm') is not None:
        cats = sorted(st.session_state['sm']['Cat'].dropna().unique())
    elif df_st_raw is not None and 'Catégorie' in df_st_raw.columns:
        cats = sorted(df_st_raw['Catégorie'].dropna().unique())
    else:
        cats = []

    st.markdown('Éditez les paramètres par catégorie. Ces paramètres sont persistés dans `parametres_stock.json`.')

    # build editable dataframe
    rows = []
    for c in cats:
        p = params.get(c, {})
        rows.append({
            'Catégorie': c,
            'lead_time_mois': int(p.get('lead_time_mois', DEFAULT_PARAMS_CATEGORIE['lead_time_mois'])),
            'stock_securite': float(p.get('stock_securite', DEFAULT_PARAMS_CATEGORIE['stock_securite'])),
            'seuil_rupture': float(p.get('seuil_rupture', DEFAULT_PARAMS_CATEGORIE['seuil_rupture'])),
            'stock_cible': float(p.get('stock_cible', DEFAULT_PARAMS_CATEGORIE['stock_cible']))
        })

    if not rows:
        st.info('Aucune catégorie détectée — générez d\'abord l\'analyse ou importez un stock.')
    else:
        try:
            df_params = st.data_editor(pd.DataFrame(rows), num_rows='dynamic', key="params_data_editor")
        except Exception:
            # fallback simple editor
            st.warning('Éditeur interactif non disponible — utilisez les champs ci-dessous.')
            edited = []
            for r in rows:
                with st.expander(r['Catégorie']):
                    lt = st.number_input(f"Lead time (mois) — {r['Catégorie']}", min_value=0, value=r['lead_time_mois'], key=f"lt_{r['Catégorie']}")
                    ss = st.number_input(f"Stock sécurité — {r['Catégorie']}", min_value=0.0, value=r['stock_securite'], key=f"ss_{r['Catégorie']}")
                    sr = st.number_input(f"Seuil rupture — {r['Catégorie']}", min_value=0.0, value=r['seuil_rupture'], key=f"sr_{r['Catégorie']}")
                    sc = st.number_input(f"Stock cible — {r['Catégorie']}", min_value=0.0, value=r['stock_cible'], key=f"sc_{r['Catégorie']}")
                    edited.append({'Catégorie': r['Catégorie'], 'lead_time_mois': int(lt), 'stock_securite': float(ss), 'seuil_rupture': float(sr), 'stock_cible': float(sc)})
            df_params = pd.DataFrame(edited)

        c1, c2 = st.columns([1, 1])
        if c1.button('Enregistrer'):
            out = {}
            for _, r in df_params.iterrows():
                out[str(r['Catégorie'])] = {
                    'lead_time_mois': int(r['lead_time_mois']),
                    'stock_securite': float(r['stock_securite']),
                    'seuil_rupture': float(r['seuil_rupture']),
                    'stock_cible': float(r['stock_cible'])
                }
            try:
                with open(params_file, 'w', encoding='utf-8') as f:
                    json.dump(out, f, ensure_ascii=False, indent=2)
                st.success('Paramètres enregistrés.')
                st.session_state['parametres_stock'] = out
            except Exception as e:
                st.error(f"Erreur lors de l'enregistrement : {e}")

        if c2.button('Réinitialiser'):
            if os.path.exists(params_file):
                try:
                    os.remove(params_file)
                except Exception:
                    st.error('Impossible de supprimer le fichier de paramètres.')
            st.session_state.pop('parametres_stock', None)
            st.success('Paramètres réinitialisés.')
    st.stop()

if page == "📚 Historique":
    st.markdown("<h2 class='woodmat-page-title'>Historique des analyses</h2>", unsafe_allow_html=True)
    hist = pd.DataFrame(st.session_state.get("analysis_history", []),
                        columns=["Date", "Utilisateur", "Catégorie", "Nombre d'articles", "Durée d'analyse"])
    st.markdown("<div class='woodmat-panel'>", unsafe_allow_html=True)
    st.dataframe(hist, use_container_width=True, height=420)
    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()

if page == "📈 Analyses":
    st.markdown("<h2 class='woodmat-page-title'>Analyses</h2>", unsafe_allow_html=True)

    # Basic checks and helpful debug messages if data is missing
    if sm is None or sm.empty:
        st.warning("Données d'analyse manquantes — générez d'abord l'analyse en important le stock actuel (et mouvements si nécessaire).")
        st.stop()

    df_mv = st.session_state.get('df_mv')
    date_max = st.session_state.get('date_max')
    date_12m_debut = st.session_state.get('date_12m_debut')

    if df_mv is None or df_mv.empty:
        st.warning("La base de mouvements est vide — impossible de calculer les analyses ABC/XYZ.")
        st.stop()

    st.markdown("<div class='woodmat-muted'>Module d'analyses métier — ABC, XYZ, Matrice, Pareto, TOP/FLOP, par catégorie, évolution et synthèse.</div>", unsafe_allow_html=True)

    # --- Data audit: show available columns to drive analyses ---
    df_st_raw = st.session_state.get('df_st_raw')
    cols_sm = list(sm.columns)
    cols_mv = list(df_mv.columns) if df_mv is not None else []
    cols_st = list(df_st_raw.columns) if df_st_raw is not None else []
    with st.expander("🔎 Audit des colonnes disponibles", expanded=False):
        st.write({'sm_columns': cols_sm, 'df_mv_columns': cols_mv, 'df_st_columns': cols_st})

    # helper
    def has(cols, name):
        return name in cols

    # Détection d'un prix unitaire réellement exploitable — jamais inventé.
    # Une colonne au nom évocateur (PU/PRIX/PRICE) ne suffit pas : elle doit aussi
    # contenir des valeurs numériques non nulles sur une part significative des
    # lignes, sinon on reste sur le proxy quantité (Sorties 12M).
    price_available = False
    price_series = None
    price_col_used = None
    if df_st_raw is not None:
        price_cols = [c for c in df_st_raw.columns if any(k in str(c).upper() for k in ['PU', 'PRIX', 'PRICE'])]
        for pc in price_cols:
            try:
                vals = pd.to_numeric(df_st_raw[pc], errors='coerce')
                if vals.notna().mean() > 0.5 and (vals.fillna(0) > 0).mean() > 0.1:
                    df_st_raw['_PU'] = vals
                    price_series = df_st_raw.groupby('Référence')['_PU'].median()
                    price_available = True
                    price_col_used = pc
                    break
            except Exception:
                continue

    # Prepare ana dataframe using only existing columns
    ana = sm.copy()
    # Standardize keys we will use
    if 'S_12M' in ana.columns:
        ana['Sorties_12M'] = ana['S_12M'].fillna(0).astype(float)
    else:
        ana['Sorties_12M'] = 0.0

    if 'S_4M' in ana.columns:
        ana['Sorties_4M'] = ana['S_4M'].fillna(0).astype(float)

    # Value proxy: do NOT invent prices — if absent, use Sorties_12M as proxy and document limitation
    if price_available and price_series is not None:
        ana = ana.set_index('Référence')
        ana['PU_detected'] = price_series
        ana['Valeur_12M'] = ana['Sorties_12M'] * ana['PU_detected'].fillna(0)
        ana = ana.reset_index()
        st.info(f"Prix détecté dans le stock actuel (colonne « {price_col_used} ») — la Valeur 12M est calculée à partir des PU détectés.")
        value_used = 'financial'
    else:
        ana['PU_detected'] = float('nan')
        ana['Valeur_12M'] = ana['Sorties_12M']
        st.info('⚠️ Limitation : aucun prix unitaire fiable détecté. L\'ABC et le Pareto utilisent Sorties 12M (quantité) comme proxy de valeur.')
        value_used = 'quantity_proxy'

    # tabs
    tab_abc, tab_xyz, tab_matrix, tab_pareto, tab_topflop, tab_cat, tab_evo, tab_synth = st.tabs([
        "Analyse ABC", "Analyse XYZ", "Matrice ABC/XYZ", "Pareto", "TOP / FLOP", "Par catégorie", "Évolution mensuelle", "Dashboard synthétique"
    ])

    # ---------- ABC ----------
    with tab_abc:
        st.markdown("### Analyse ABC")
        # ensure required display columns exist
        for c in ['Référence', 'Designation', 'Cat']:
            if c not in ana.columns:
                ana[c] = 'N/A'

        df_abc = ana[['Référence', 'Designation', 'Cat', 'Sorties_12M', 'Valeur_12M']].copy()
        df_abc = df_abc.sort_values('Valeur_12M', ascending=False).reset_index(drop=True)
        total_value = float(df_abc['Valeur_12M'].sum())
        df_abc['Cumul'] = df_abc['Valeur_12M'].cumsum()
        if total_value > 0:
            df_abc['Cumul_pct'] = df_abc['Cumul'] / total_value
        else:
            df_abc['Cumul_pct'] = 0.0
        def abc_class(v):
            if v <= 0.80:
                return 'A'
            if v <= 0.95:
                return 'B'
            return 'C'
        if total_value == 0:
            df_abc['Classe_ABC'] = 'C'
            st.warning("Aucune sortie sur 12M détectée — l'ABC est indisponible par valeur; toutes les références sont marquées 'C'.")
        else:
            df_abc['Classe_ABC'] = df_abc['Cumul_pct'].apply(abc_class)

        counts = df_abc['Classe_ABC'].value_counts().reindex(['A','B','C']).fillna(0).astype(int)
        pct_refs = (counts / len(df_abc) * 100).round(1) if len(df_abc) > 0 else counts
        val_pct = df_abc.groupby('Classe_ABC')['Valeur_12M'].sum().reindex(['A','B','C']).fillna(0)
        val_pct = (val_pct / (total_value if total_value > 0 else 1) * 100).round(1)
        st.markdown(f"Références total : **{len(df_abc)}**")
        st.write(pd.DataFrame({
            'Nb références': counts,
            '% références': pct_refs,
            '% valeur': val_pct
        }))
        add_export_buttons(df_abc, 'analyse_abc_table', 'ABC détaillé', date_max)
        st.dataframe(df_abc.rename(columns={'Designation':'Désignation','Cat':'Catégorie','Sorties_12M':'Sorties 12M','Valeur_12M':'Valeur 12M'}), use_container_width=True, height=420)
        # Pareto chart (uses Valeur_12M which may be proxy)
        fig = go.Figure()
        fig.add_trace(go.Bar(x=df_abc['Référence'], y=df_abc['Valeur_12M'], name='Valeur 12M'))
        fig.add_trace(go.Scatter(x=df_abc['Référence'], y=(df_abc['Cumul_pct']*100), name='Cumul %', yaxis='y2'))
        fig.update_layout(yaxis2=dict(overlaying='y', side='right', range=[0,100]), height=420, margin=dict(l=10,r=10,t=20,b=120))
        st.plotly_chart(fig, use_container_width=True)

    # ---------- XYZ ----------
    with tab_xyz:
        st.markdown("### Analyse XYZ — régularité de la demande")
        s_all = df_mv[df_mv['ES']=='S'] if 'ES' in df_mv.columns else pd.DataFrame()
        df_xyz = pd.DataFrame(columns=['Référence', 'Moy_Mois', 'Std_Mois', 'CV', 'Classe_XYZ'])
        if s_all.empty:
            st.warning('Aucune sortie enregistrée dans les mouvements — XYZ indisponible.')
        else:
            # monthly series for last 12 months
            s_12m = s_all[(s_all['Date'] >= date_12m_debut) & (s_all['Date'] <= date_max)].copy()
            s_12m['Mois'] = s_12m['Date'].dt.to_period('M').dt.to_timestamp()
            piv = s_12m.pivot_table(index='Reference', columns='Mois', values='Qty', aggfunc='sum', fill_value=0)
            # ensure 12 months columns present
            months = pd.date_range(date_12m_debut.normalize(), periods=12, freq='MS')
            for m in months:
                if m not in piv.columns:
                    piv[m] = 0
            piv = piv[sorted(piv.columns)]
            stats = []
            for ref, row in piv.iterrows():
                vals = row.values.astype(float)
                mean = float(vals.mean())
                std = float(vals.std(ddof=0))
                if mean == 0:
                    cv = float('inf')
                else:
                    cv = std / mean
                stats.append((ref, mean, std, cv))
            df_xyz = pd.DataFrame(stats, columns=['Référence', 'Moy_Mois', 'Std_Mois', 'CV'])
            # classify
            def xyz_label(r):
                if r['Moy_Mois'] == 0:
                    return 'Z / Sans demande'
                cv = r['CV']
                if cv <= 0.5:
                    return 'X'
                if cv <= 1.0:
                    return 'Y'
                return 'Z'
            df_xyz['Classe_XYZ'] = df_xyz.apply(xyz_label, axis=1)
            counts_xyz = df_xyz['Classe_XYZ'].value_counts().reindex(['X','Y','Z','Z / Sans demande']).fillna(0).astype(int)
            st.write(pd.DataFrame({'Nb références': counts_xyz}))
            merge_cols = [c for c in ['Référence','Designation','Cat','S_12M','S_4M'] if c in ana.columns]
            df_xyz = df_xyz.merge(ana[merge_cols], on='Référence', how='left')
            add_export_buttons(df_xyz, 'analyse_xyz_table', 'XYZ détaillé', date_max)
            st.dataframe(df_xyz.sort_values(['Classe_XYZ','CV']), use_container_width=True, height=420)

    # ---------- ABC x XYZ matrix ----------
    with tab_matrix:
        st.markdown("### Matrice ABC × XYZ")
        if df_abc.empty or df_xyz.empty:
            st.warning('Données ABC ou XYZ insuffisantes pour générer la matrice — importez des mouvements de sortie.')
            merged = pd.DataFrame(columns=['Référence', 'Classe_ABC', 'Classe_XYZ'])
        else:
            merged = df_abc[['Référence','Classe_ABC']].merge(df_xyz[['Référence','Classe_XYZ']], on='Référence', how='inner')
            matrix = pd.crosstab(merged['Classe_ABC'], merged['Classe_XYZ'])
            st.dataframe(matrix, use_container_width=True)
            interpretations = {
                'AX': 'Priorité maximale, demande importante et régulière',
                'AY': 'Importante mais variable',
                'AZ': 'Importante mais très irrégulière, gestion prudente',
                'BX': 'Gestion standard', 'BY': 'Gestion standard', 'BZ': 'Surveillance',
                'CX': 'Faible priorité', 'CY': 'Faible priorité', 'CZ': 'Faible valeur et demande irrégulière / potentiellement dormant'
            }
            st.markdown("**Interprétation automatique (exemples)**")
            for k,v in interpretations.items():
                st.markdown(f"- **{k}** → {v}")

    # ---------- Pareto ----------
    with tab_pareto:
        st.markdown("### Pareto — valeur cumulée")
        cats = sorted(ana['Cat'].dropna().unique()) if 'Cat' in ana.columns else []
        choice = st.selectbox("Filtrer par catégorie", options=['Toutes'] + cats, key="analyse_pareto_categorie")
        dfp = ana.copy()
        if choice != 'Toutes':
            dfp = dfp[dfp['Cat'] == choice]
        dfp = dfp.sort_values('Valeur_12M', ascending=False).reset_index(drop=True)
        dfp['Cumul'] = dfp['Valeur_12M'].cumsum()
        total = float(dfp['Valeur_12M'].sum())
        dfp['Cumul_pct'] = dfp['Cumul'] / (total if total > 0 else 1)
        if total == 0:
            st.warning('Aucune valeur disponible pour le Pareto (Sorties 12M = 0).')
        fig = go.Figure()
        fig.add_trace(go.Bar(x=dfp['Référence'], y=dfp['Valeur_12M'], name='Valeur'))
        fig.add_trace(go.Scatter(x=dfp['Référence'], y=dfp['Cumul_pct']*100, name='Cumul %', yaxis='y2'))
        fig.add_hline(y=80, line_dash='dash', line_color='green')
        fig.add_hline(y=95, line_dash='dash', line_color='orange')
        fig.update_layout(yaxis2=dict(overlaying='y', side='right', range=[0,100]), height=420, margin=dict(b=120))
        st.plotly_chart(fig, use_container_width=True)
        export_cols = [c for c in ['Référence','S_12M','Valeur_12M','Cumul_pct'] if c in dfp.columns]
        add_export_buttons(dfp[export_cols], 'pareto', 'Pareto', date_max)

    # ---------- TOP / FLOP ----------
    with tab_topflop:
        st.markdown("### TOP / FLOP — synthèse")
        # TOP lists - use only existing columns
        top_s12_cols = [c for c in ['Référence','Designation','Cat','S_12M','S_4M','Rotation_Actuelle','Rotation_12M'] if c in ana.columns]
        top_s12 = ana.sort_values('Sorties_12M', ascending=False).head(10)[[c for c in top_s12_cols if c in ana.columns]]
        top_s4 = ana.sort_values('Sorties_4M', ascending=False).head(10)[[c for c in top_s12_cols if c in ana.columns and 'S_4M' in ana.columns]] if 'S_4M' in ana.columns else pd.DataFrame()
        top_rot = ana.sort_values('Rotation_Actuelle', ascending=False).head(10)[[c for c in ['Référence','Designation','Cat','Rotation_Actuelle','Sorties_12M','Stock'] if c in ana.columns]] if 'Rotation_Actuelle' in ana.columns else pd.DataFrame()
        st.markdown('#### 🔥 TOP références (Sorties 12M)')
        st.dataframe(top_s12, use_container_width=True)
        st.markdown('#### 🔥 TOP références (Sorties 4M)')
        st.dataframe(top_s4, use_container_width=True)
        st.markdown('#### 🔥 TOP références (Rotation actuelle)')
        st.dataframe(top_rot, use_container_width=True)

        # ── FLOP — score multi-critères ──────────────────────────────────────
        # Ne repose plus uniquement sur la Rotation. Chaque critère DISPONIBLE
        # contribue un score 0-1 (rang percentile, sens "plus mauvais = plus
        # proche de 1") ; les critères indisponibles sont exclus du score (pas
        # inventés) et listés comme limitation.
        flop = ana.copy()
        criteres_utilises = []
        criteres_absents = []
        score_cols = []

        def _pct_rank(s, ascending):
            # rang percentile 0..1 ; ascending=True -> les valeurs FAIBLES obtiennent un score élevé
            r = s.rank(pct=True, method='average')
            return (1 - r) if ascending else r

        if 'Sorties_12M' in flop.columns:
            flop['Score_faible_sortie'] = _pct_rank(flop['Sorties_12M'], ascending=True)
            score_cols.append('Score_faible_sortie')
            criteres_utilises.append('Faible ou aucune sortie (12M)')
        else:
            criteres_absents.append('Sorties 12M')

        if 'Stock' in flop.columns:
            flop['Score_stock_eleve'] = _pct_rank(flop['Stock'], ascending=False)
            score_cols.append('Score_stock_eleve')
            criteres_utilises.append('Stock élevé')
        else:
            criteres_absents.append('Stock')

        if 'Couverture' in flop.columns:
            flop['Score_couverture_elevee'] = _pct_rank(flop['Couverture'].fillna(0), ascending=False)
            score_cols.append('Score_couverture_elevee')
            criteres_utilises.append('Couverture élevée')
        else:
            criteres_absents.append('Couverture')

        if 'Dern_Sortie' in flop.columns and date_max is not None:
            _jours_depuis = (pd.Timestamp(date_max) - pd.to_datetime(flop['Dern_Sortie'], errors='coerce')).dt.days
            flop['Jours_depuis_derniere_sortie'] = _jours_depuis
            flop['Score_derniere_sortie_ancienne'] = _pct_rank(_jours_depuis.fillna(_jours_depuis.max() if _jours_depuis.notna().any() else 0), ascending=False)
            score_cols.append('Score_derniere_sortie_ancienne')
            criteres_utilises.append('Dernière sortie ancienne')
        else:
            criteres_absents.append('Date de dernière sortie')

        if 'Taux_Immob' in flop.columns:
            flop['Score_immobilisation'] = _pct_rank(flop['Taux_Immob'], ascending=False)
            score_cols.append('Score_immobilisation')
            criteres_utilises.append("Taux d'immobilisation élevé")
        else:
            criteres_absents.append("Taux d'immobilisation")

        if not score_cols:
            st.warning('Aucune colonne disponible pour calculer un score FLOP.')
        else:
            flop['Score_FLOP'] = flop[score_cols].mean(axis=1).round(3)

            def _classe_flop(v):
                if v >= 0.80:
                    return '🔴 FLOP critique'
                if v >= 0.60:
                    return '🟠 FLOP probable'
                return '🟡 À surveiller'

            flop_cols_display = [c for c in ['Référence','Designation','Cat','Sorties_12M','Stock','Couverture',
                                              'Taux_Immob','Dern_Sortie','Jours_depuis_derniere_sortie','Score_FLOP'] if c in flop.columns]
            flop_top = flop.sort_values('Score_FLOP', ascending=False).head(50).copy()
            flop_top['Classe_FLOP'] = flop_top['Score_FLOP'].apply(_classe_flop)
            st.markdown('#### ⚠️ FLOP — score multi-critères')
            st.caption(f"Critères utilisés : {', '.join(criteres_utilises)}." +
                       (f" Non disponibles (exclus, jamais inventés) : {', '.join(criteres_absents)}." if criteres_absents else ""))
            st.dataframe(flop_top[flop_cols_display + ['Classe_FLOP']], use_container_width=True)
            add_export_buttons(flop_top[flop_cols_display + ['Classe_FLOP']], 'analyse_flop', 'FLOP', date_max)

    # ---------- Par catégorie ----------
    with tab_cat:
        st.markdown('### Analyse par catégorie')
        st.caption("La rotation globale par catégorie est un ratio d'agrégats (ΣSorties ÷ ΣStock), "
                   "jamais une moyenne des rotations individuelles des références.")
        # build aggregation using available columns; missing columns will produce NaN but table will show availability
        agg_map = {
            'Nb_ref': ('Référence','nunique'),
            'Stock_total': ('Stock','sum'),
            'Sorties_12M': ('Sorties_12M','sum'),
            'Moy_Mois_12M': ('Moy_Mois_12M','mean'),
            'Sorties_4M': ('Sorties_4M','sum'),
            'Moy_Mois_4M': ('Moy_Mois_4M','mean'),
            'Stock_moyen_12M_total': ('Stock_moyen_12M', 'sum'),
            'Stock_moyen_4M_total': ('Stock_moyen_4M', 'sum'),
            'Couverture_moy': ('Couverture','mean'),
            'Immobilisation_moy': ('Taux_Immob','mean'),
            'Dormant': ('Class', lambda s: (s=='Dormant').sum()),
            'A_risque_rupture': ('Couverture', lambda s: (s <= 1).sum())
        }
        valid_aggs = {k:v for k,v in agg_map.items() if v[0] in ana.columns}
        if not valid_aggs:
            st.warning('Colonnes nécessaires à l\'analyse par catégorie manquantes.')
        else:
            grp = ana.groupby('Cat').agg(**{k: v for k,v in valid_aggs.items()}).reset_index()
            # Les 3 rotations par catégorie sont TOUJOURS des ratios d'agrégats
            # (Σ Sorties ÷ Σ Stock de la catégorie), jamais une moyenne des rotations
            # individuelles des références — cohérent avec le KPI Dashboard.
            if {'Stock_total', 'Sorties_12M'}.issubset(grp.columns):
                grp['Rotation_actuelle_cat'] = 0.0
                m_grp = grp['Stock_total'] > SEUIL
                grp.loc[m_grp, 'Rotation_actuelle_cat'] = (grp.loc[m_grp, 'Sorties_12M'] / grp.loc[m_grp, 'Stock_total']).round(2)
            if {'Stock_moyen_12M_total', 'Sorties_12M'}.issubset(grp.columns):
                grp['Rotation_12M_historique_cat'] = 0.0
                m_h12 = grp['Stock_moyen_12M_total'] > SEUIL
                grp.loc[m_h12, 'Rotation_12M_historique_cat'] = (grp.loc[m_h12, 'Sorties_12M'] / grp.loc[m_h12, 'Stock_moyen_12M_total']).round(2)
            if {'Stock_moyen_4M_total', 'Sorties_4M'}.issubset(grp.columns):
                grp['Rotation_4M_cat'] = 0.0
                m_h4 = grp['Stock_moyen_4M_total'] > SEUIL
                grp.loc[m_h4, 'Rotation_4M_cat'] = (grp.loc[m_h4, 'Sorties_4M'] / grp.loc[m_h4, 'Stock_moyen_4M_total']).round(2)
            st.dataframe(grp, use_container_width=True)
            st.caption("Rotation actuelle = Σ Sorties 12M ÷ Σ Stock actuel de la catégorie. "
                       "Rotation 12M historique = Σ Sorties 12M ÷ Σ Stock moyen 12M de la catégorie. "
                       "Rotation 4M = Σ Sorties 4M ÷ Σ Stock moyen 4M de la catégorie. "
                       "Les 3 sont des ratios d'agrégats, jamais des moyennes de rotations individuelles.")
            add_export_buttons(grp, 'analyse_par_categorie', 'Par catégorie', date_max)

    # ---------- Évolution mensuelle ----------
    with tab_evo:
        st.markdown('### Évolution mensuelle (12 derniers mois)')
        refs_filter = st.selectbox('Filtrer par référence (optionnel)', options=['Toutes'] + sorted(ana['Référence'].astype(str).unique()) if 'Référence' in ana.columns else ['Toutes'], key="analyse_evolution_reference")
        cats_filter = st.selectbox('Filtrer par catégorie', options=['Toutes'] + sorted(ana['Cat'].dropna().unique()) if 'Cat' in ana.columns else ['Toutes'], key="analyse_evolution_categorie")
        evo_df = evolution_mensuelle(df_mv, None)
        if refs_filter != 'Toutes':
            evo_df = evolution_mensuelle(df_mv, [refs_filter])
        if cats_filter != 'Toutes':
            refs_in_cat = ana[ana['Cat']==cats_filter]['Référence'].tolist() if 'Cat' in ana.columns else []
            evo_df = evolution_mensuelle(df_mv, refs_in_cat)
        if evo_df.empty:
            st.warning('Pas de sorties mensuelles disponibles pour la période sélectionnée.')
        else:
            evo_df = evo_df.sort_values('Mois')
            evo_df['MA3'] = evo_df['Qty'].rolling(3, min_periods=1).mean()
            fig = go.Figure()
            fig.add_trace(go.Bar(x=evo_df['Mois'], y=evo_df['Qty'], name='Sorties mensuelles'))
            fig.add_trace(go.Line(x=evo_df['Mois'], y=evo_df['MA3'], name='Moyenne mobile 3M'))
            st.plotly_chart(fig, use_container_width=True)
            add_export_buttons(evo_df, 'evolution_mensuelle', 'Évolution mensuelle', date_max)

    # ---------- Dashboard synthétique ----------
    with tab_synth:
        st.markdown('### Dashboard synthétique')
        total_refs = ana['Référence'].nunique() if 'Référence' in ana.columns else 0
        n_abc_a = df_abc[df_abc['Classe_ABC']=='A']['Référence'].nunique() if not df_abc.empty else 0
        n_xyz_z = df_xyz[df_xyz['Classe_XYZ'].astype(str).str.startswith('Z')]['Référence'].nunique() if not df_xyz.empty else 0
        n_ax = merged[(merged['Classe_ABC']=='A') & (merged['Classe_XYZ'].astype(str).str.startswith('X'))]['Référence'].nunique() if 'merged' in locals() and not merged.empty else 0
        n_rupture = int((ana['Class']=='Rupture').sum()) if 'Class' in ana.columns else 0
        n_dormant = int((ana['Class']=='Dormant').sum()) if 'Class' in ana.columns else 0

        # « Références sous seuil » : ne dépend plus d'une valeur fixe de 1 mois.
        # Utilise le Seuil de rupture RÉEL de chaque catégorie (⚙️ Paramètres /
        # parametres_stock.json), via le même moteur que Réapprovisionnement/Alertes.
        _params_synth = charger_parametres_stock()
        if {'Cat', 'Stock'}.issubset(ana.columns):
            _reappro_synth = calculer_moteur_reappro(ana, _params_synth)
            n_sous_seuil = int(_reappro_synth['Risque_Reappro'].astype(str).str.startswith('🔴').sum())
            seuil_label = "Sous seuil de rupture (par catégorie)"
        else:
            n_sous_seuil = 0
            seuil_label = "Sous seuil de rupture (non calculable)"

        st.write({'Total références': total_refs, 'ABC-A': n_abc_a, 'XYZ-Z': n_xyz_z, 'AX': n_ax,
                   'Rupture': n_rupture, seuil_label: n_sous_seuil, 'Dormant': n_dormant})
        st.caption("« Sous seuil » utilise désormais le Seuil de rupture propre à chaque catégorie "
                   "(⚙️ Paramètres), et non plus une valeur fixe de 1 mois pour toutes les catégories.")

        # Stock immobilisé : AUCUNE valeur financière (pas de prix disponible → "Non disponible").
        # Mesure physique uniquement, et jamais une somme brute entre unités différentes
        # (m³, m², P, ML additionnés n'a pas de sens) : volume m³ + nb de références dormantes.
        st.markdown('#### 🧱 Stock immobilisé (mesure physique)')
        st.info("💰 Valeur financière du stock immobilisé : **Non disponible** — aucun prix unitaire fiable "
                "dans les données actuelles.")
        if {'Stock', 'Unite'}.issubset(ana.columns):
            vol_m3 = ana.loc[ana['Unite'].astype(str).str.upper().eq('M3'), 'Stock'].sum()
            dormant_refs = ana[ana['Class'] == 'Dormant'] if 'Class' in ana.columns else ana.iloc[0:0]
            dormant_vol_m3 = dormant_refs.loc[dormant_refs['Unite'].astype(str).str.upper().eq('M3'), 'Stock'].sum() if not dormant_refs.empty else 0.0
            ic1, ic2, ic3 = st.columns(3)
            ic1.metric('Volume total en stock (m³)', format_nombre_fr(vol_m3, 2))
            ic2.metric('Références dormantes', f"{len(dormant_refs):,}".replace(',', ' '))
            ic3.metric('Volume dormant (m³)', format_nombre_fr(dormant_vol_m3, 2))
            st.caption("Le stock exprimé dans d'autres unités (m², P, ML) n'est pas additionné au volume m³ "
                       "— unités non convertibles entre elles. Détail par unité ci-dessous.")
            st.dataframe(ana.groupby('Unite', as_index=False)['Stock'].sum().rename(
                columns={'Unite': 'Unité', 'Stock': 'Stock total (unité native)'}), use_container_width=True)
        else:
            st.warning("Colonnes Stock / Unité manquantes — mesure physique du stock immobilisé non disponible.")
        st.markdown('#### TOP 10 (Sorties 12M)')
        if 'Sorties_12M' in ana.columns:
            st.dataframe(ana.sort_values('Sorties_12M', ascending=False).head(10)[[c for c in ['Référence','Designation','Cat','Sorties_12M','Stock'] if c in ana.columns]], use_container_width=True)
        else:
            st.info('Sorties 12M non disponibles — TOP non calculable.')
        st.markdown('#### FLOP 10 (score multi-critères)')
        if 'flop_top' in locals() and not flop_top.empty:
            st.dataframe(flop_top[flop_cols_display + ['Classe_FLOP']].head(10), use_container_width=True)
        else:
            st.info('Pas de FLOP calculable avec les données disponibles.')

    st.success('Analyses générées à partir des données chargées.')

if sm is None:
    st.markdown(f"<h2 class='woodmat-page-title'>{page}</h2>", unsafe_allow_html=True)
    st.info("⬅️ Importez le **stock actuel** (et les mouvements de l'année en cours si disponibles) "
            "dans la barre latérale, puis cliquez sur **Générer l'analyse**.")
    st.stop()

df_mv = st.session_state["df_mv"]
date_max = st.session_state["date_max"]
date_12m_debut = st.session_state["date_12m_debut"]

st.caption(
    f"Analyse réalisée le : {date_max.strftime('%d/%m/%Y')}  |  "
    f"Fenêtre d'analyse : 12 et 4 derniers mois  |  "
    f"Mouvements analysés : {len(df_mv):,}".replace(',', ' ')
)

# ── Filtres ──────────────────────────────────────────────
fc1, fc2, fc3 = st.columns([2, 2, 2])
fc4, fc5 = st.columns([2, 2])
with fc1:
    cats = sorted(sm['Cat'].dropna().unique())
    sel_cats = st.multiselect("Catégorie", cats, default=[], key="filtre_principal_categorie")
with fc2:
    classes = sorted(sm['Class'].unique())
    sel_class = st.multiselect("Classification", classes, default=[], key="filtre_principal_classification")
with fc3:
    unites = sorted(sm['Unite'].dropna().astype(str).unique())
    sel_unites = st.multiselect("Unité", unites, default=[], key="filtre_principal_unite")
with fc4:
    references = sorted(sm['Référence'].dropna().astype(str).unique())
    sel_refs = st.multiselect("Référence", references, default=[], key="filtre_principal_reference")
with fc5:
    recherche = st.text_input("🔍 Recherche référence ou désignation", key="filtre_principal_recherche")

f = sm.copy()
if sel_cats:
    f = f[f['Cat'].isin(sel_cats)]
if sel_class:
    f = f[f['Class'].isin(sel_class)]
if sel_unites:
    f = f[f['Unite'].astype(str).isin(sel_unites)]
if sel_refs:
    f = f[f['Référence'].astype(str).isin(sel_refs)]
if recherche:
    f = f[f['Référence'].astype(str).str.contains(recherche, case=False, na=False)
          | f['Designation'].astype(str).str.contains(recherche, case=False, na=False)]

CLASS_BADGE = {
    'Rupture': '🔴 Rupture', 'Critique': '🔴 Critique', 'Stock faible': '🟠 Stock faible',
    'Normal': '🟡 Normal', 'Bon niveau': '🟢 Bon niveau', 'Surstock': '🔵 Surstock',
    'Surstock important': '⚫ Surstock important / Dormant', 'Dormant': '⚫ Dormant',
    'Sans mouvement': '📦 Sans mouvement',
}

def badge_class(series):
    return series.map(lambda v: CLASS_BADGE.get(v, v))

st.markdown(f"<h2 class='woodmat-page-title'>{page}</h2>", unsafe_allow_html=True)

display_cols = ['Référence', 'Designation', 'Cat', 'Unite', 'Stock', 'Class',
                 'S_12M', 'Moy_Mois_12M', 'S_4M', 'Moy_Mois_4M',
                 'Rotation_Actuelle', 'Rotation_12M', 'Rotation_4M', 'Tendance_Label',
                 'Couverture', 'Taux_Immob', 'Dern_Sortie']
rename_cols = {'Designation': 'Désignation', 'Cat': 'Catégorie', 'Unite': 'Unité', 'Class': 'Classification',
               'S_12M': 'Sorties 12M', 'Moy_Mois_12M': 'Moy/Mois 12M',
               'S_4M': 'Sorties 4M', 'Moy_Mois_4M': 'Moy/Mois 4M',
               'Rotation_Actuelle': 'Rotation actuelle', 'Rotation_12M': 'Rotation 12M historique',
               'Rotation_4M': 'Rotation 4M',
               'Tendance_Label': 'Tendance 4M vs 12M',
               'Couverture': 'Couv. (mois)', 'Taux_Immob': 'Immob. (%)',
               'Dern_Sortie': 'Dern. Sortie'}

if page == "📦 Réapprovisionnement":
    st.caption("Aide à la décision basée sur les paramètres par catégorie (Délai d'approvisionnement, "
               "Seuil de rupture, Stock de sécurité, Stock cible) définis dans ⚙️ Paramètres. Le CUMP n'est pas utilisé.")
    _params = charger_parametres_stock()
    if not _params:
        st.warning("Aucun paramètre enregistré dans ⚙️ Paramètres — les valeurs par défaut sont utilisées "
                   f"(Délai {DEFAULT_PARAMS_CATEGORIE['lead_time_mois']:.0f} mois, "
                   f"Seuil de rupture {DEFAULT_PARAMS_CATEGORIE['seuil_rupture']:.0f} mois, "
                   f"Stock de sécurité {DEFAULT_PARAMS_CATEGORIE['stock_securite']:.0f} mois, "
                   f"Stock cible {DEFAULT_PARAMS_CATEGORIE['stock_cible']:.0f} mois).")
    rep = calculer_moteur_reappro(f, _params)

    kc1, kc2, kc3, kc4 = st.columns(4)
    filters = {
        '🔴 Risque de rupture': rep[rep['Risque_Reappro'].str.startswith('🔴')],
        '⚠️ Commander maintenant': rep[rep['Risque_Reappro'].str.startswith('⚠️')],
        '🟢 Couverture suffisante': rep[rep['Risque_Reappro'].str.startswith('🟢')],
        '⚪ Sans demande récente': rep[rep['Risque_Reappro'].str.startswith('⚪')],
    }
    for col, label in zip([kc1, kc2, kc3, kc4], filters):
        if col.button(f"{label}\n\n{len(filters[label])} article(s)", use_container_width=True):
            st.session_state['reappro_filter'] = label
    selected_priority = st.session_state.get('reappro_filter')
    if selected_priority:
        st.info(f"Filtre actif : {selected_priority}")
        rep = filters[selected_priority]

    rep_cols = ['Référence', 'Designation', 'Cat', 'Unite', 'Stock', 'Conso_Mensuelle_Moyenne',
                'Couverture_Reappro', 'Delai_Appro_Mois', 'Seuil_Rupture_Mois', 'Stock_Securite_Mois',
                'Stock_Cible_Mois', 'Sous_Stock_Securite', 'Risque_Reappro', 'Qte_Recommandee', 'Fiabilite']
    rep_df = rep[[c for c in rep_cols if c in rep.columns]].rename(columns={
        'Designation': 'Désignation', 'Cat': 'Catégorie', 'Unite': 'Unité', 'Stock': 'Stock actuel',
        'Conso_Mensuelle_Moyenne': 'Conso. mensuelle moy. (4M réels)', 'Couverture_Reappro': 'Couverture (mois)',
        'Delai_Appro_Mois': "Délai d'appro (mois)", 'Seuil_Rupture_Mois': 'Seuil de rupture (mois)',
        'Stock_Securite_Mois': 'Stock de sécurité (mois)', 'Stock_Cible_Mois': 'Stock cible (mois)',
        'Sous_Stock_Securite': 'Sous le stock de sécurité', 'Risque_Reappro': 'Risque',
        'Qte_Recommandee': 'Quantité recommandée', 'Fiabilite': 'Fiabilité',
    })
    add_export_buttons(rep_df, 'reapprovisionnement', 'Réapprovisionnement', date_max)
    st.dataframe(rep_df.sort_values('Couverture (mois)', na_position='first'), use_container_width=True, height=520)
    st.caption("Quantité recommandée = max(0, Stock cible × Conso. mensuelle moyenne − Stock actuel). "
               "Marquée « Non fiable » quand aucune sortie n'a eu lieu sur les 4 derniers mois — "
               "la consommation moyenne n'est alors pas calculable à partir de données réelles.")
    st.stop()

if page == "😴 Stock dormant":
    st.caption("Articles classés Dormant uniquement — recherche, filtres et exports dédiés.")
    dorm = f[f['Class'] == 'Dormant'][display_cols].rename(columns=rename_cols)
    dorm['Classification'] = badge_class(dorm['Classification'])
    add_export_buttons(dorm, 'stock_dormant', 'Stock Dormant', date_max)
    st.dataframe(dorm.sort_values('Immob. (%)', ascending=False), use_container_width=True, height=520)
    st.stop()

if page == "🪵 Stock Bois Rouge":
    st.caption("Articles BOIS ROUGE uniquement — recherche, filtres et exports dédiés.")
    br = f[f['Cat'].astype(str).str.upper().eq('BOIS ROUGE')][display_cols].rename(columns=rename_cols)
    br['Classification'] = badge_class(br['Classification'])
    add_export_buttons(br, 'stock_bois_rouge_articles', 'Articles Bois Rouge', date_max)
    st.dataframe(br, use_container_width=True, height=520)
    st.divider()
    df_st_raw = st.session_state.get("df_st_raw")
    if df_st_raw is not None:
        br_bytes, non_reconnus = generer_excel_bois_rouge(df_st_raw, date_max)
        if br_bytes:
            st.download_button("⬇️ Télécharger la synthèse Qualité × Fournisseur × Dimension", br_bytes,
                               file_name=f"stock_bois_rouge_synthese_{date_max.strftime('%d_%m_%Y')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
    st.stop()

if page == "⚠️ Alertes":
    st.caption("Articles en rupture et articles dormants, complétés par le risque de rupture calculé "
               "à partir du Seuil de rupture et du Délai d'approvisionnement par catégorie (⚙️ Paramètres).")
    _params_al = charger_parametres_stock()
    f_risque = calculer_moteur_reappro(f, _params_al)
    alert_df = f_risque[
        f_risque['Class'].isin(['Rupture', 'Critique', 'Dormant', 'Sans mouvement'])
        | f_risque['Risque_Reappro'].str.startswith('🔴')
        | f_risque['Risque_Reappro'].str.startswith('⚠️')
    ]
    alert_cols = ['Référence', 'Designation', 'Cat', 'Unite', 'Stock', 'Class', 'Risque_Reappro',
                  'S_12M', 'S_4M', 'Dern_Sortie']
    alert_df = alert_df[[c for c in alert_cols if c in alert_df.columns]].rename(columns={
        'Designation': 'Désignation', 'Cat': 'Catégorie', 'Unite': 'Unité',
        'Class': 'Classification', 'Risque_Reappro': 'Risque de rupture (paramétré)',
        'S_12M': 'Sorties 12M', 'S_4M': 'Sorties 4M', 'Dern_Sortie': 'Dern. Sortie'
    })
    if 'Classification' in alert_df.columns:
        alert_df['Classification'] = badge_class(alert_df['Classification']) if 'badge_class' in globals() else alert_df['Classification']
    add_export_buttons(alert_df, 'alertes', 'Alertes', date_max)
    st.dataframe(alert_df, use_container_width=True, height=520)
    st.stop()

if page == "📄 Rapports":
    st.caption("Exports disponibles pour la vue filtrée courante.")
    report_cols = ['Référence', 'Designation', 'Cat', 'Unite', 'Stock', 'Class', 'S_12M', 'Moy_Mois_12M',
                   'S_4M', 'Moy_Mois_4M', 'Rotation_Actuelle', 'Rotation_12M', 'Rotation_4M', 'Tendance_Label']
    report_df = f[report_cols].rename(columns=rename_cols)

    # Résumé de la vue courante
    total_refs = len(report_df)
    total_stock = report_df['Stock'].sum() if 'Stock' in report_df.columns else 0.0
    total_sorties_12m = report_df['Sorties 12M'].sum() if 'Sorties 12M' in report_df.columns else 0.0
    rotation_globale = total_sorties_12m / total_stock if total_stock > 0 else float('nan')
    c1, c2, c3, c4 = st.columns(4)
    c1.metric('Références', f'{total_refs:,}'.replace(',', ' '))
    c2.metric('Stock total', format_nombre_fr(total_stock, 3))
    c3.metric('Sorties 12M', format_nombre_fr(total_sorties_12m, 2))
    c4.metric('Rotation globale', format_nombre_fr(rotation_globale, 2))

    st.markdown('#### Filtres actifs')
    active_filters = {
        'Catégorie': sel_cats if sel_cats else 'Toutes',
        'Classification': sel_class if sel_class else 'Toutes',
        'Unité': sel_unites if sel_unites else 'Toutes',
        'Références': f'{len(sel_refs)} sélectionnée(s)' if sel_refs else 'Toutes',
        'Recherche': recherche or 'Aucune'
    }
    st.write(active_filters)

    st.divider()
    add_export_buttons(report_df, 'rapport_rotation', 'Rapport Rotation', date_max)
    st.dataframe(report_df, use_container_width=True, height=420)

    # Historique des analyses
    hist = st.session_state.get('analysis_history', [])
    if hist:
        with st.expander('Historique des analyses récentes', expanded=False):
            hist_df = pd.DataFrame(hist)
            st.dataframe(hist_df.sort_values('Date', ascending=False).reset_index(drop=True), use_container_width=True, height=320)
    else:
        st.info('Aucune analyse historique disponible pour l\'instant.')
    st.stop()

# ── KPIs ─────────────────────────────────────────────────
stock_par_unite = (
    f.assign(Unite_Affichage=f['Unite'].apply(format_unite_stock))
     .groupby('Unite_Affichage', as_index=False)['Stock']
     .sum()
)
unite_order = {'m³': 0, 'm²': 1, 'P': 2, 'ML': 3}
stock_par_unite['_Ordre'] = stock_par_unite['Unite_Affichage'].map(unite_order).fillna(99)
stock_par_unite = stock_par_unite.sort_values(['_Ordre', 'Unite_Affichage'])
volume_stock_html = "<br>".join(
    f"{format_nombre_fr(row.Stock)} {row.Unite_Affichage}"
    for row in stock_par_unite.itertuples(index=False)
) or "—"
# KPI historique — formule d'origine INCHANGÉE (continuité avec les anciens rapports)
# Calcul retenu (Dashboard "Rotation du stock") :
#   Rotation globale = (Somme des Sorties 12M) ÷ (Somme du Stock ACTUEL)
# Justification : le KPI Dashboard doit rester la vue agrégée "Sorties 12M ÷ Stock ACTUEL".
# Remarque importante : on inclut ici l'ensemble des articles disposant d'un stock
# significatif (> SEUIL) pour éviter d'exclure involontairement des références
# dont le stock actuel est non nul (exclusion précédente sur 'Rotation_12M>0' pouvait
# retirer des articles avec sorties mais stock moyen recalculé à zéro, faussant le ratio).
# Ne pas confondre avec les KPI analytiques `rot_globale_12m` (stock moyen) ou
# `Rotation_12M` / `Rotation_Actuelle` individuels qui restent calculés séparément.
_base_rot = f[f['Stock'] > SEUIL]
rot_moy = _base_rot['S_12M'].sum() / _base_rot['Stock'].sum() if _base_rot['Stock'].sum() > 0 else float('nan')

# Nouveau KPI, séparé : Rotation globale 12M sur stock moyen global (ne remplace pas rot_moy)
rot_globale_12m = _base_rot['S_12M'].sum() / _base_rot['Stock_moyen_12M'].sum() if _base_rot['Stock_moyen_12M'].sum() > 0 else float('nan')

# Nouveau KPI, séparé : Rotation globale 4M sur stock moyen global 4M (jamais mélangé au 12M)
_base_rot4 = f[f['Rotation_4M'] > 0]
rot_globale_4m = _base_rot4['S_4M'].sum() / _base_rot4['Stock_moyen_4M'].sum() if _base_rot4['Stock_moyen_4M'].sum() > 0 else float('nan')

n_rupture = len(f[f['Class'] == 'Rupture'])
n_critique = len(f[f['Class'] == 'Critique'])
n_dormant = len(f[f['Class'].isin(['Dormant', 'Sans mouvement'])])
n_alertes = n_rupture + n_critique + n_dormant

k1, k2, k3, k4 = st.columns(4)
with k1:
    st.markdown(
        f"<div class='woodmat-kpi-card'><div class='woodmat-kpi-title'>Volume Stock</div>"
        f"<div class='woodmat-kpi-value'>{volume_stock_html}</div></div>",
        unsafe_allow_html=True)
with k2:
    st.metric(
        "Rotation du stock",
        f"{rot_moy:.2f} tours/an" if pd.notna(rot_moy) else "—",
        help="Sorties des 12 derniers mois ÷ Stock ACTUEL total de la catégorie. KPI de référence du Dashboard — inchangé.")
    st.markdown("<div class='woodmat-muted'>Calcul sur les 12 derniers mois — stock actuel</div>", unsafe_allow_html=True)

# ── Audit temporaire du KPI "Rotation du stock" ──────────
# Permet de vérifier à l'écran les totaux exacts utilisés par le calcul et les
# filtres actifs, pour comprendre pourquoi la valeur affichée est celle qu'elle est.
with st.expander("🔎 Audit — Rotation du stock (diagnostic temporaire)", expanded=False):
    st.write({
        'Total Sorties 12M (Σ S_12M, Stock > seuil)': format_nombre_fr(_base_rot['S_12M'].sum(), 3),
        'Total Stock actuel (Σ Stock, Stock > seuil)': format_nombre_fr(_base_rot['Stock'].sum(), 3),
        'Rotation calculée (Σ Sorties 12M ÷ Σ Stock actuel)': f"{rot_moy:.2f}" if pd.notna(rot_moy) else "—",
        'Nb références incluses (Stock > seuil)': int(len(_base_rot)),
        'Nb références filtrées (total vue courante)': int(len(f)),
        'Filtres actifs': {
            'Catégorie': sel_cats if sel_cats else 'Toutes',
            'Classification': sel_class if sel_class else 'Toutes',
            'Unité': sel_unites if sel_unites else 'Toutes',
            'Référence': f'{len(sel_refs)} sélectionnée(s)' if sel_refs else 'Toutes',
            'Recherche': recherche or 'Aucune',
        },
    })
with k3:
    st.metric(
        "⚠️ Alertes",
        "Articles nécessitant une action",
        help="Regroupe les ruptures de stock, les articles en couverture critique (<1 mois) et les articles sans sortie récente (Dormant / Sans mouvement).")
    st.markdown(
        f"<div class='woodmat-kpi-detail'>Rupture : {n_rupture}<br>"
        f"Critique : {n_critique}<br>"
        f"Dormant / Sans mouvement : {n_dormant}<br><strong>Total : {n_alertes}</strong></div>",
        unsafe_allow_html=True)
with k4:
    st.metric("Articles analysés (références)", len(f))

st.divider()

# ── Graphique évolution des ventes ──────────────────────
gcol, tcol = st.columns([1.3, 1])
with gcol:
    st.markdown("<div class='woodmat-section-title'>Historique des sorties mensuelles</div>", unsafe_allow_html=True)
    st.markdown("<div class='woodmat-muted'>Quantités sorties par mois sur les 12 derniers mois.</div>", unsafe_allow_html=True)
    refs_filtrees = f['Référence'].tolist() if (sel_cats or sel_class or sel_unites or sel_refs or recherche) else None
    evo = evolution_mensuelle(df_mv, refs_filtrees)
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=evo['Mois'], y=evo['Qty'], mode='lines+markers',
                              line=dict(color='#1F3864', width=2), fill='tozeroy',
                              fillcolor='rgba(31,56,100,0.1)'))
    fig.update_layout(height=340, margin=dict(l=10, r=10, t=10, b=10),
                       xaxis_title=None, yaxis_title="Quantité sortie")
    st.plotly_chart(fig, use_container_width=True)

with tcol:
    st.markdown("<div class='woodmat-section-title'>Répartition des articles par niveau de rotation</div>", unsafe_allow_html=True)
    dist = f['Class'].value_counts().reset_index()
    dist.columns = ['Classification', 'Nb']
    fig2 = px.pie(dist, names='Classification', values='Nb', hole=0.45,
                  color='Classification', color_discrete_map=CLASS_COLORS)
    fig2.update_layout(height=340, margin=dict(l=10, r=10, t=10, b=10), showlegend=True)
    st.plotly_chart(fig2, use_container_width=True)
    st.markdown("<div class='woodmat-muted'>Classification calculée selon la rotation observée sur les 12 derniers mois.</div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div class='woodmat-legend'>
        <strong>Rupture</strong> → Stock nul.<br>
        <strong>Critique</strong> → Couverture &lt; 1 mois face à la demande récente (4M).<br>
        <strong>Stock faible</strong> → Couverture entre 1 et 3 mois.<br>
        <strong>Normal</strong> → Couverture entre 3 et 6 mois.<br>
        <strong>Bon niveau</strong> → Couverture entre 6 et 12 mois.<br>
        <strong>Surstock</strong> → Couverture entre 12 et 18 mois.<br>
        <strong>Surstock important / Dormant</strong> → Couverture ≥ 18 mois.<br>
        <strong>Dormant</strong> → Déjà vendu par le passé mais aucune sortie sur les 4 derniers mois.<br>
        <strong>Sans mouvement</strong> → Aucune sortie enregistrée depuis l'origine de l'historique.
        </div>
        """,
        unsafe_allow_html=True)

st.divider()

# ── Tableau dynamique ───────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs(["📋 Tableau complet", "😴 Stock dormant", "📅 Historique par année",
                                    "🪵 Stock Bois Rouge"])

with tab1:
    tdf = f[display_cols].rename(columns=rename_cols).sort_values('Rotation actuelle', ascending=False)
    tdf['Classification'] = badge_class(tdf['Classification'])
    st.dataframe(tdf, use_container_width=True, height=480)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        tdf.to_excel(writer, index=False, sheet_name='Rotation Stock')
    st.download_button("⬇️ Exporter Excel (vue actuelle)", buf.getvalue(),
                        file_name=f"rotation_{date_max.strftime('%d_%m_%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tab2:
    dorm = f[f['Class'] == 'Dormant'][display_cols].rename(columns=rename_cols)
    dorm['Classification'] = badge_class(dorm['Classification'])
    st.caption(f"{len(dorm)} référence(s) dormante(s) — aucune sortie sur les 4 derniers mois "
               f"mais historique de sorties existant.")
    st.dataframe(dorm.sort_values('Immob. (%)', ascending=False),
                 use_container_width=True, height=420)
    buf2 = io.BytesIO()
    with pd.ExcelWriter(buf2, engine='openpyxl') as writer:
        dorm.to_excel(writer, index=False, sheet_name='Stock Dormant')
    st.download_button("⬇️ Exporter Stock Dormant", buf2.getvalue(),
                        file_name=f"stock_dormant_{date_max.strftime('%d_%m_%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tab3:
    cols_annuelles = st.session_state.get("cols_annuelles", [])
    s_cols = sorted([c for c in cols_annuelles if c.startswith('S_')])
    a_cols = sorted([c for c in cols_annuelles if c.startswith('A_')])
    t_cols = sorted([c for c in cols_annuelles if c.startswith('T_')])

    st.caption("Détail des sorties (S), achats/entrées (A) et nombre de transactions (T) "
               "par référence, année par année — même filtres que le tableau principal.")

    vue = st.radio("Vue", ["Sorties par année", "Achats/Entrées par année", "Nb transactions par année"],
                    horizontal=True, key="historique_vue_annee")
    if vue == "Sorties par année":
        cols_show = s_cols
    elif vue == "Achats/Entrées par année":
        cols_show = a_cols
    else:
        cols_show = t_cols

    hist_cols = ['Référence', 'Designation', 'Cat', 'Unite', 'Stock', 'Class'] + cols_show
    hist_cols = [c for c in hist_cols if c in f.columns]
    hdf = f[hist_cols].rename(columns={'Designation': 'Désignation', 'Cat': 'Catégorie', 'Unite': 'Unité', 'Class': 'Classification'})
    if 'Classification' in hdf.columns:
        hdf['Classification'] = badge_class(hdf['Classification'])
    st.dataframe(hdf, use_container_width=True, height=480)

    buf3 = io.BytesIO()
    with pd.ExcelWriter(buf3, engine='openpyxl') as writer:
        hdf.to_excel(writer, index=False, sheet_name='Historique par annee')
    st.download_button("⬇️ Exporter cette vue", buf3.getvalue(),
                        file_name=f"historique_annuel_{date_max.strftime('%d_%m_%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tab4:
    df_st_raw = st.session_state.get("df_st_raw")
    if df_st_raw is None or 'Catégorie' not in df_st_raw.columns:
        st.info("Générez d'abord l'analyse (fichier stock actuel) pour voir cet onglet.")
    else:
        st.caption("Stock BOIS ROUGE + BOIS BLANC SUEDE (même famille ENSO/STORA ENSO) par Qualité × "
                   "Fournisseur × Dimension — même mise en forme "
                   "que l'ancien fichier Excel. **ENSO** et **STORA ENSO** sont fusionnés (même fournisseur).")
        br_bytes, non_reconnus = generer_excel_bois_rouge(df_st_raw, date_max)
        if br_bytes is None:
            st.warning("Aucune référence BOIS ROUGE exploitable trouvée dans le stock actuel.")
        else:
            if len(non_reconnus) > 0:
                total_absent = non_reconnus['Quantité'].sum() if 'Quantité' in non_reconnus.columns else 0
                st.warning(f"⚠️ {len(non_reconnus)} référence(s) BOIS ROUGE avec un fournisseur non "
                           f"reconnu ({total_absent:.3f} M3) ne figurent pas dans le tableau ci-dessous. "
                           f"Dites-moi les noms de fournisseurs manquants et je les ajoute.")
            st.download_button("⬇️ Télécharger la feuille Stock Bois Rouge (Excel)", br_bytes,
                                file_name=f"stock_bois_rouge_{date_max.strftime('%d_%m_%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary")
